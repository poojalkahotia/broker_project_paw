# invapp/views/party_views.py

from django.shortcuts import render, get_object_or_404, redirect
from brokerapp.forms import PartyForm, BrokerForm, ItemForm, FirmForm 
from brokerapp.models import HeadParty, Broker, HeadItem ,SaleMaster, SaleDetails ,PurchaseMaster, PurchaseDetails, DailyPage, JamaEntry, NaameEntry, Firm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Max
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from django.http import JsonResponse
import json
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
from django.utils import timezone
from django.db.models import Sum, Prefetch, F, FloatField, ExpressionWrapper
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.db.models import ProtectedError
from io import BytesIO
from fpdf import FPDF
from django.views.generic import TemplateView
from .forms import AllPartyBalanceForm
from django.urls import reverse
from urllib.parse import quote
import io
from decimal import Decimal, InvalidOperation
from django.core.mail import send_mail
from django.core.mail import EmailMessage

from django.conf import settings

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    get_column_letter = None

class PDF(FPDF):
    def header(self):
        self.set_font("Arial", "B", 14)
        self.cell(0, 10, "Daily Report", ln=1, align="C")
        self.ln(5)

    def table_side(self, title, entries, total, x_pos, max_rows):
        """Side table banane ke liye"""
        start_y = self.get_y()
        self.set_xy(x_pos, start_y)
        self.set_font("Arial", "B", 12)
        self.cell(90, 10, title, ln=1, align="L")

        # header
        self.set_x(x_pos)
        self.set_font("Arial", "B", 10)
        self.cell(20, 8, "No", 1, 0, "C")
        self.cell(40, 8, "Party", 1, 0, "C")
        self.cell(30, 8, "Amount", 1, 1, "C")

        # rows
        self.set_font("Arial", "", 10)
        for e in entries:
            self.set_x(x_pos)
            self.cell(20, 8, str(e.entry_no), 1, 0, "C")
            self.cell(40, 8, e.party.partyname, 1, 0)
            self.cell(30, 8, f"{e.amount:.2f}", 1, 1, "R")

        # padding blank rows
        extra_rows = max_rows - len(entries)
        for _ in range(extra_rows):
            self.set_x(x_pos)
            self.cell(20, 8, "", 1, 0, "C")
            self.cell(40, 8, "", 1, 0)
            self.cell(30, 8, "", 1, 1, "R")

        # total row
        self.set_x(x_pos)
        self.set_font("Arial", "B", 10)
        self.cell(60, 8, "Total", 1, 0, "R")
        self.cell(30, 8, f"{total:.2f}", 1, 1, "R")



# -----------------------
# Helpers
# -----------------------
def to_decimal(val, default=Decimal('0')):
    """Safe conversion to Decimal."""
    try:
        return Decimal(str(val))
    except (InvalidOperation, TypeError, ValueError):
        return default




# -----------------------
# Views
# -----------------------

def sale_form(request, invno=None):
    """
    Render sale form. If invno provided, load sale and its details (scoped to current org).
    """
    sale = None
    sale_items_json = "[]"
    today_date = date.today().strftime("%Y-%m-%d")

    if invno:
        # sale must belong to current org
        sale = get_object_or_404(SaleMaster, invno=invno, org=request.current_org)
        details = SaleDetails.objects.filter(salemaster=sale)
        items_data = []
        for d in details:
            items_data.append({
                "item_id": d.item.pk,
                "item_name": d.item.item_name,
                "bora": float(d.bora),
                "bn": float(d.bn),
                "bnwt": float(d.bnwt),
                "bo": float(d.bo),
                "bowt": float(d.bowt),
                "tbwt": float(getattr(d, "tbwt", 0)),
                "totalbora": float(d.bn*d.bnwt + d.bo*d.bowt),
                "qty": float(d.qty),
                "rate": float(d.rate),
                "amt": float(d.amount),
                "partywt": float(d.partywt),
                "millwt": float(d.millwt),
                "diffwt": float(d.diffwt),
                "frkwt": float(getattr(d, "frkwt", 0)),
                "lotno": d.lotno or "",
            })
        sale_items_json = json.dumps(items_data)

    # next invoice number — per ORG
    next_invno = SaleMaster.objects.filter(org=request.current_org).aggregate(Max("invno"))['invno__max']
    next_invno = (next_invno + 1) if next_invno else 1

    # Firms: try to scope to current_org if Firm has org FK, otherwise return all firms
    try:
        # if Firm has 'org' field this will work; otherwise it raises FieldDoesNotExist or similar
        firms_qs = Firm.objects.filter(org=request.current_org).order_by('firmname')
    except Exception:
        firms_qs = Firm.objects.all().order_by('firmname')
    
    context = {
        "sale": sale,
        "sale_items_json": sale_items_json,
        "next_invno": next_invno,
        "today_date": today_date,
        # only current org choices
        "items": HeadItem.objects.filter(org=request.current_org).order_by('item_name'),
        "parties": HeadParty.objects.filter(org=request.current_org).order_by('partyname'),
        "brokers": Broker.objects.filter(org=request.current_org).order_by('brokername'),
        "firms": firms_qs,
    }
    return render(request, "brokerapp/sale.html", context)

# ===================== SAVE SALE =====================


@transaction.atomic
def save_sale(request):
    """
    Save a new SaleMaster and its SaleDetails — scoped to current org.
    """
    if request.method != "POST":
        return redirect("sale_form_new")

    try:
        invdate_str = request.POST.get("invdate")
        invdate = datetime.strptime(invdate_str, "%Y-%m-%d").date() if invdate_str else date.today()
        awakno = request.POST.get("awakno", "").strip()
        extra = request.POST.get("extra", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        firm_pk = request.POST.get("firm")
        vehicleno = request.POST.get("vehicleno", "").strip()

        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("sale_form_new")

        total_amt = Decimal("0")
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal("100")).quantize(Decimal("0.01"))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal("100")).quantize(Decimal("0.01"))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))
        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal("0.01"))
        netamt = (total - advance).quantize(Decimal("0.01"))

        # Resolve FKs inside same org
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        # Firm resolve
        firm = None
        if firm_pk:
            firm_fields = [f.name for f in Firm._meta.get_fields()]
            if "org" in firm_fields:
                firm = get_object_or_404(Firm, pk=firm_pk, org=request.current_org)
            else:
                firm = get_object_or_404(Firm, pk=firm_pk)

        # Create SaleMaster
        sale = SaleMaster.objects.create(
            org=request.current_org,
            created_by=request.user,
            invdate=invdate,
            awakno=awakno,
            party=party,
            broker=broker,
            firm=firm,
            vehicleno=vehicleno,
            extra=extra,
            totalamt=total_amt.quantize(Decimal("0.01")),
            batavpercent=batavpercent,
            batavamt=batavamt,
            dr=dr,
            dramt=dramt,
            qi=qi,
            other=other,
            total=total,
            advance=advance,
            netamt=netamt,
            remark=request.POST.get("remark", "").strip(),
        )

        # Save Items
        for it in items:
            item_id = it.get("item_id")
            item_obj = get_object_or_404(HeadItem, pk=item_id, org=request.current_org)
            SaleDetails.objects.create(
                salemaster=sale,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                tbwt=to_decimal(it.get("tbwt", 0)),
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                frkwt=to_decimal(it.get("frkwt", 0)),
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        
        # -------------------------
        #  EMAIL + PROFESSIONAL PDF INVOICE
        # -------------------------
        send_email = request.POST.get("send_email") == "on"

        if send_email and party.email:
            # FPDF installed nahin hua to simple warning
            if FPDF is None:
                messages.warning(
                    request,
                    "Invoice email ke liye 'fpdf' package chahiye. Install: pip install fpdf"
                )
            else:
                # helper: UTF-8 -> ascii only
                def safe_text(val, maxlen=None):
                    s = "" if val is None else str(val)
                    s = s.replace("—", "-").replace("–", "-")
                    s = "".join(ch if ord(ch) < 128 else "?" for ch in s)
                    return s[:maxlen] if maxlen else s

                try:
                    # ---------- 1) Professional looking invoice PDF ----------
                    pdf = FPDF("P", "mm", "A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.set_margins(15, 15, 15)
                    pdf.add_page()

                    # Top org name
                    org_name = getattr(sale.org, "name", str(sale.org))
                    pdf.set_font("Helvetica", "B", 16)
                    pdf.cell(0, 8, safe_text(org_name, 60), ln=1)

                    pdf.set_font("Helvetica", "", 10)
                    pdf.cell(0, 6, safe_text("Sale Invoice", 40), ln=1)
                    pdf.ln(2)

                    # Invoice meta (left-right panel)
                    pdf.set_font("Helvetica", "", 10)

                    pdf.cell(28, 6, "Invoice No:", 0, 0)
                    pdf.cell(60, 6, safe_text(sale.invno), 0, 0)
                    pdf.cell(20, 6, "Date:", 0, 0)
                    pdf.cell(0, 6, safe_text(sale.invdate.strftime("%d-%m-%Y")), 0, 1)

                    pdf.cell(28, 6, "Party:", 0, 0)
                    pdf.cell(60, 6, safe_text(sale.party.partyname, 40), 0, 0)
                    pdf.cell(20, 6, "Vehicle:", 0, 0)
                    pdf.cell(0, 6, safe_text(sale.vehicleno or "-", 30), 0, 1)

                    pdf.cell(28, 6, "Firm:", 0, 0)
                    firm_label = getattr(sale.firm, "firmname", "") if sale.firm else ""
                    pdf.cell(60, 6, safe_text(firm_label, 40), 0, 0)
                    pdf.cell(20, 6, "Broker:", 0, 0)
                    pdf.cell(0, 6, safe_text(sale.broker.brokername, 40), 0, 1)

                    pdf.ln(4)

                    # Items table
                    headers = ["Item", "Bora", "TBwt", "Qty", "FrkWt", "Rate", "Amount", "LotNo"]
                    # total width ~ 180mm (A4 width 210 - margins 15+15)
                    widths  = [45,   15,    18,    15,    18,     22,    25,      22]

                    pdf.set_font("Helvetica", "B", 9)
                    for i, h in enumerate(headers):
                        pdf.cell(widths[i], 7, safe_text(h, 20), border=1, align="C")
                    pdf.ln(7)

                    pdf.set_font("Helvetica", "", 9)
                    details_qs = SaleDetails.objects.filter(salemaster=sale).select_related("item")

                    for d in details_qs:
                        row = [
                            safe_text(d.item.item_name, 30),
                            safe_text(f"{d.bora:.0f}", 5),
                            safe_text(f"{d.tbwt:.2f}", 10),
                            safe_text(f"{d.qty:.2f}", 10),
                            safe_text(f"{d.frkwt:.2f}", 10),
                            safe_text(f"{d.rate:.2f}", 10),
                            safe_text(f"{d.amount:.2f}", 12),
                            safe_text(d.lotno or "", 12),
                        ]
                        for i, v in enumerate(row):
                            align = "R" if i in (1, 2, 3, 4, 5, 6) else "L"
                            pdf.cell(widths[i], 7, v, border=1, align=align)
                        pdf.ln(7)

                    pdf.ln(3)

                    # Totals block on right side
                    pdf.set_font("Helvetica", "", 9)
                    right_x = pdf.w - 15 - 60   # 60mm wide totals box
                    pdf.set_xy(right_x, pdf.get_y())

                    lines = [
                        ("Total Amount",    sale.totalamt),
                        (f"Batav ({sale.batavpercent:.2f}%)", sale.batavamt),
                        (f"Dr ({sale.dr:.2f}%)", sale.dramt),
                        ("QI + Other",      sale.qi + sale.other),
                        ("Advance",         sale.advance),
                        ("Net Amount",      sale.netamt),
                    ]
                    for label, val in lines:
                        pdf.cell(35, 6, safe_text(label, 30), border=0, align="R")
                        pdf.cell(25, 6, f"{val:.2f}", border=0, align="R")
                        pdf.ln(6)

                    # footer note
                    pdf.ln(4)
                    pdf.set_x(15)
                    pdf.set_font("Helvetica", "I", 8)
                    pdf.cell(0, 5, "This is a system generated invoice.", 0, 1)

                    buf = io.BytesIO()
                    pdf.output(buf)
                    buf.seek(0)
                    pdf_bytes = buf.read()

                    safe_party = "".join(
                        ch if ord(ch) < 128 else "?" for ch in party.partyname
                    )[:40]

                    # ---------- 2) Email with attached PDF ----------
                    subject = f"Sale Invoice #{sale.invno}"
                    body = (
                        f"Dear {party.partyname},\n\n"
                        "Please find your sale invoice attached as PDF.\n\n"
                        f"Total Amount: {sale.totalamt:.2f}\n"
                        f"Net Amount  : {sale.netamt:.2f}\n\n"
                        "Thank you."
                    )

                    email = EmailMessage(
                        subject,
                        body,
                        settings.DEFAULT_FROM_EMAIL,
                        [party.email],
                    )
                    email.attach(
                        f"invoice_{sale.invno}_{safe_party}.pdf",
                        pdf_bytes,
                        "application/pdf",
                    )
                    email.send(fail_silently=False)

                except Exception as e:
                    messages.warning(request, f"Sale saved but invoice email not sent: {e}")

        messages.success(request, "Sale entry saved successfully!")
        return redirect("saledata")
        

    except Exception as e:
        messages.error(request, f"Error saving sale: {e}")
        return redirect("sale_form_new")


@transaction.atomic
def update_sale(request, invno):
    sale = get_object_or_404(SaleMaster, invno=invno, org=request.current_org)

    if request.method != "POST":
        return redirect('sale_form_update', invno=invno)

    try:
        # ---------- HEADER FIELDS ----------
        invdate_str = request.POST.get("invdate")
        invdate = datetime.strptime(invdate_str, "%Y-%m-%d").date() if invdate_str else sale.invdate
        awakno = request.POST.get("awakno", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        vehicleno = request.POST.get("vehicleno", "").strip()
        extra = request.POST.get("extra", "").strip()

        # ---------- ITEMS ----------
        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("sale_form_update", invno=invno)

        total_amt = Decimal("0")
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal('100')).quantize(Decimal("0.01"))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal('100')).quantize(Decimal("0.01"))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))

        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal("0.01"))
        netamt = (total - advance).quantize(Decimal("0.01"))

        # ---------- FOREIGN KEYS ----------
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        firm = None
        firm_pk = request.POST.get("firm")
        if firm_pk:
            firm_fields = [f.name for f in Firm._meta.get_fields()]
            if "org" in firm_fields:
                firm = get_object_or_404(Firm, pk=firm_pk, org=request.current_org)
            else:
                firm = get_object_or_404(Firm, pk=firm_pk)

        # ---------- UPDATE HEADER ----------
        sale.invdate = invdate
        sale.awakno = awakno
        sale.party = party
        sale.broker = broker
        sale.vehicleno = vehicleno
        sale.extra = extra
        sale.totalamt = total_amt.quantize(Decimal("0.01"))
        sale.batavpercent = batavpercent
        sale.batavamt = batavamt
        sale.dr = dr
        sale.dramt = dramt
        sale.qi = qi
        sale.other = other
        sale.advance = advance
        sale.total = total
        sale.netamt = netamt
        sale.remark = request.POST.get("remark", "").strip()
        if firm:
            sale.firm = firm
        sale.save()

        # ---------- REPLACE ITEM ROWS ----------
        SaleDetails.objects.filter(salemaster=sale).delete()
        for it in items:
            item_obj = get_object_or_404(HeadItem, pk=it.get("item_id"), org=request.current_org)
            SaleDetails.objects.create(
                salemaster=sale,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                tbwt=to_decimal(it.get("tbwt", 0)),
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                frkwt=to_decimal(it.get("frkwt", 0)),
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        # --------------------------------------------------------------------
        # EMAIL + PDF (PROFESSIONAL INVOICE) — CORRECT POSITION, CLEAN INDENT
        # --------------------------------------------------------------------
        send_email = request.POST.get("send_email") == "on"

        if send_email and party.email:
            if FPDF is None:
                messages.warning(request, "PDF library missing: pip install fpdf")
            else:

                def safe_text(v, maxlen=None):
                    s = "" if v is None else str(v)
                    s = "".join(ch if ord(ch) < 128 else "?" for ch in s)
                    return s[:maxlen] if maxlen else s

                try:
                    # Make PDF (same code as save_sale)
                    pdf = FPDF("P", "mm", "A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.set_margins(15, 15, 15)
                    pdf.add_page()

                    # Header
                    org_name = getattr(sale.org, "name", "")
                    pdf.set_font("Helvetica", "B", 16)
                    pdf.cell(0, 8, safe_text(org_name), ln=1)

                    pdf.set_font("Helvetica", "", 10)
                    pdf.cell(0, 6, "Sale Invoice (Updated)", ln=1)
                    pdf.ln(2)

                    # Details
                    pdf.set_font("Helvetica", "", 10)
                    pdf.cell(28, 6, "Invoice No:", 0, 0)
                    pdf.cell(60, 6, safe_text(sale.invno), 0, 0)
                    pdf.cell(20, 6, "Date:", 0, 0)
                    pdf.cell(0, 6, sale.invdate.strftime("%d-%m-%Y"), 0, 1)

                    pdf.cell(28, 6, "Party:", 0, 0)
                    pdf.cell(60, 6, safe_text(party.partyname), 0, 0)
                    pdf.cell(20, 6, "Vehicle:", 0, 0)
                    pdf.cell(0, 6, safe_text(sale.vehicleno or "-"), 0, 1)

                    pdf.ln(4)

                    # Table
                    headers = ["Item", "Bora", "TBwt", "Qty", "FrkWt", "Rate", "Amount", "LotNo"]
                    widths  = [45, 15, 18, 15, 18, 22, 25, 22]

                    pdf.set_font("Helvetica", "B", 9)
                    for i, h in enumerate(headers):
                        pdf.cell(widths[i], 7, safe_text(h), border=1, align="C")
                    pdf.ln(7)

                    pdf.set_font("Helvetica", "", 9)
                    rows = SaleDetails.objects.filter(salemaster=sale)

                    for d in rows:
                        row = [
                            safe_text(d.item.item_name),
                            f"{d.bora:.0f}",
                            f"{d.tbwt:.2f}",
                            f"{d.qty:.2f}",
                            f"{d.frkwt:.2f}",
                            f"{d.rate:.2f}",
                            f"{d.amount:.2f}",
                            safe_text(d.lotno),
                        ]
                        for i, v in enumerate(row):
                            pdf.cell(widths[i], 7, v, border=1, align="R" if i>0 else "L")
                        pdf.ln(7)

                    # Totals
                    pdf.ln(3)
                    right_x = pdf.w - 15 - 60
                    pdf.set_xy(right_x, pdf.get_y())

                    totals = [
                        ("Total Amount", total_amt),
                        (f"Batav ({batavpercent:.2f}%)", batavamt),
                        (f"Dr ({dr:.2f}%)", dramt),
                        ("QI + Other", qi + other),
                        ("Advance", advance),
                        ("Net Amount", netamt),
                    ]

                    for lbl, val in totals:
                        pdf.cell(35, 6, safe_text(lbl), 0, 0, "R")
                        pdf.cell(25, 6, f"{val:.2f}", 0, 1, "R")

                    # Footer
                    pdf.ln(4)
                    pdf.set_x(15)
                    pdf.set_font("Helvetica", "I", 8)
                    pdf.cell(0, 5, "This is a system-generated invoice.")

                    # Output
                    buf = io.BytesIO()
                    pdf.output(buf)
                    buf.seek(0)
                    pdf_bytes = buf.read()

                    safe_party = "".join(ch if ord(ch)<128 else "?" for ch in party.partyname)

                    subject = f"Updated Sale Invoice #{sale.invno}"
                    body = (
                        f"Dear {party.partyname},\n\n"
                        "Your sale invoice has been UPDATED.\n\n"
                        f"Total Amount: {total_amt:.2f}\n"
                        f"Net Amount  : {netamt:.2f}\n\n"
                        "Please find attached updated invoice PDF.\n\n"
                        "Thank you."
                    )

                    email = EmailMessage(
                        subject,
                        body,
                        settings.DEFAULT_FROM_EMAIL,
                        [party.email],
                    )
                    email.attach(
                        f"invoice_{sale.invno}_{safe_party}.pdf",
                        pdf_bytes,
                        "application/pdf"
                    )
                    email.send()

                except Exception as e:
                    messages.warning(request, f"Invoice email not sent: {e}")

        messages.success(request, "Sale entry updated successfully!")
        return redirect("saledata")

    except Exception as e:
        messages.error(request, f"Error updating sale: {e}")
        return redirect("sale_form_update", invno=invno)

def sale_data_view(request):
    """List of sales (scoped to current org)."""
    sales = SaleMaster.objects.filter(org=request.current_org).order_by("-invno")
    return render(request, "brokerapp/saledata.html", {
        "sales": sales,
        "today_date": date.today(),
    })




def delete_sale(request, invno):
    sale = get_object_or_404(SaleMaster, invno=invno, org=request.current_org)
    sale.delete()
    messages.success(request, "Sale entry deleted successfully!")
    return redirect("saledata")




def sale_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    broker_id = request.GET.get("broker")
    firm_id = request.GET.get("firm")         # optional firm filter (pk or name)
    report_type = request.GET.get("report_type", "date")

    # Default date = today
    if not start_date:
        start_date = date.today().strftime("%Y-%m-%d")
    if not end_date:
        end_date = date.today().strftime("%Y-%m-%d")

    # Base queryset (ORG SCOPED) + prefetch details (with item) for the template
    sales = (
        SaleMaster.objects
        .filter(org=request.current_org)
        .select_related("broker", "firm")   # firm selected for template access
        .prefetch_related(Prefetch("details", queryset=SaleDetails.objects.select_related("item")))
    )

    if start_date:
        sales = sales.filter(invdate__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(invdate__lte=parse_date(end_date))

    # Broker filter: allow pk or name, but scoped to org
    if broker_id and broker_id != "all":
        # try as primary key
        if sales.filter(broker__pk=broker_id).exists():
            sales = sales.filter(broker__pk=broker_id)
        else:
            sales = sales.filter(broker__brokername=broker_id)

    # Firm filter: allow pk or name, scoped to org (only apply when provided and not "all")
    if firm_id and firm_id != "all":
        # try pk (firm FK)
        if sales.filter(firm__pk=firm_id).exists():
            sales = sales.filter(firm__pk=firm_id)
        else:
            sales = sales.filter(firm__firmname=firm_id)

    sales = sales.order_by("invdate")

    report_data = []

    # --- Grouping & Aggregates ---
    if report_type == "date":
        grouped = sales.values("invdate").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate")

        for g in grouped:
            group_sales = sales.filter(invdate=g["invdate"])
            # TBWt & FrkWt sum for this group (sum over details)
            tbwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_tbwt=Sum("tbwt"))["total_tbwt"] or 0
            frkwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_frkwt=Sum("frkwt"))["total_frkwt"] or 0

            g["total_tbwt"] = tbwt_sum
            g["total_frkwt"] = frkwt_sum

            report_data.append({
                "group": g["invdate"],
                "items": group_sales,
                "totals": g
            })

    elif report_type == "broker":
        grouped = sales.values("invdate", "broker__brokername").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate", "broker__brokername")

        for g in grouped:
            group_sales = sales.filter(
                invdate=g["invdate"],
                broker__brokername=g["broker__brokername"]
            )
            tbwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_tbwt=Sum("tbwt"))["total_tbwt"] or 0
            frkwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_frkwt=Sum("frkwt"))["total_frkwt"] or 0

            g["total_tbwt"] = tbwt_sum
            g["total_frkwt"] = frkwt_sum

            report_data.append({
                "group": f"{g['invdate']} - {g['broker__brokername'] or 'No Broker'}",
                "items": group_sales,
                "totals": g
            })

    elif report_type == "firm":
        # Group by invdate + firm (firm may be null)
        grouped = sales.values("invdate", "firm__firmname").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate", "firm__firmname")

        for g in grouped:
            firm_name = g.get("firm__firmname")  # may be None
            if firm_name is None:
                # sales with NULL firm for that date
                group_sales = sales.filter(invdate=g["invdate"], firm__isnull=True)
                label_firm = "No Firm"
            else:
                group_sales = sales.filter(invdate=g["invdate"], firm__firmname=firm_name)
                label_firm = firm_name

            tbwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_tbwt=Sum("tbwt"))["total_tbwt"] or 0
            frkwt_sum = SaleDetails.objects.filter(salemaster__in=group_sales).aggregate(total_frkwt=Sum("frkwt"))["total_frkwt"] or 0

            g["total_tbwt"] = tbwt_sum
            g["total_frkwt"] = frkwt_sum

            report_data.append({
                "group": f"{g['invdate']} - {label_firm}",
                "items": group_sales,
                "totals": g
            })

    # Overall Totals (header-level) + TBWt + FrkWt across all details in the filtered set
    overall_totals = sales.aggregate(
        total_totalamt=Sum("totalamt"),
        total_batavamt=Sum("batavamt"),
        total_dramt=Sum("dramt"),
        total_other=Sum("other"),
        total_total=Sum("total"),
        total_advance=Sum("advance"),
        total_netamt=Sum("netamt"),
    )
    overall_tbwt = SaleDetails.objects.filter(salemaster__in=sales).aggregate(total_tbwt=Sum("tbwt"))["total_tbwt"] or 0
    overall_frkwt = SaleDetails.objects.filter(salemaster__in=sales).aggregate(total_frkwt=Sum("frkwt"))["total_frkwt"] or 0

    overall_totals["total_tbwt"] = overall_tbwt
    overall_totals["total_frkwt"] = overall_frkwt

    # Dropdowns also ORG SCOPED
    brokers = Broker.objects.filter(org=request.current_org).order_by("brokername")
    
    try:
        firm_field_names = [f.name for f in Firm._meta.get_fields()]
        if 'org' in firm_field_names:
            firms = Firm.objects.filter(org=request.current_org).order_by('firmname')
        else:
            firms = Firm.objects.all().order_by('firmname')
    except Exception:
        # defensive fallback: return all firms if meta inspection fails
        firms = Firm.objects.all().order_by('firmname')

    context = {
        "report_data": report_data,
        "overall_totals": overall_totals,
        "start_date": start_date,
        "end_date": end_date,
        "brokers": brokers,
        "firms": firms,
        "selected_broker": broker_id if broker_id != "all" else None,
        "selected_firm": firm_id if firm_id != "all" else None,
        "report_type": report_type,
    }
    return render(request, "brokerapp/sale_report.html", context)

# ===================== PDF (FPDF) =====================
def sale_report_pdf(request):
    """
    Generate Sale Report PDF (FPDF) using current filters.
    Includes per-invoice detail rows with TBWt and FrkWt.
    Numbers are right-aligned with thousand separators.
    """
    # --- build same filtered queryset as HTML report ---
    start_date = request.GET.get("start_date") or date.today().strftime("%Y-%m-%d")
    end_date = request.GET.get("end_date") or date.today().strftime("%Y-%m-%d")
    broker_id = request.GET.get("broker")
    firm_id = request.GET.get("firm")          # NEW: firm filter
    report_type = request.GET.get("report_type", "date")

    sales = (
        SaleMaster.objects
        .filter(org=request.current_org)
        .select_related("broker", "firm")   # firm selected for template/pdf access
        .prefetch_related(Prefetch("details", queryset=SaleDetails.objects.select_related("item")))
    )
    if start_date:
        sales = sales.filter(invdate__gte=parse_date(start_date))
    if end_date:
        sales = sales.filter(invdate__lte=parse_date(end_date))

    if broker_id and broker_id != "all":
        if sales.filter(broker__pk=broker_id).exists():
            sales = sales.filter(broker__pk=broker_id)
        else:
            sales = sales.filter(broker__brokername=broker_id)

    # APPLY FIRM FILTER (same behavior as in HTML view)
    if firm_id and firm_id != "all":
        if sales.filter(firm__pk=firm_id).exists():
            sales = sales.filter(firm__pk=firm_id)
        else:
            sales = sales.filter(firm__firmname=firm_id)

    sales = sales.order_by("invdate", "invno")

    # group-key helpers (just for headings)
    if report_type == "date":
        def group_key(s): return (s.invdate,)
    elif report_type == "broker":
        def group_key(s): return (s.invdate, s.broker.brokername if s.broker else "")
    else:  # report_type == "firm" (or others) -> use firm as second key
        def group_key(s): return (s.invdate, s.firm.firmname if s.firm else "")

    # --- FPDF setup ---
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Sale Report", ln=1, align="C")
    pdf.set_font("Helvetica", "", 9)
    hdr = f"From {start_date} To {end_date} | Generated: {timezone.now().strftime('%d-%m-%Y %I:%M %p')}"
    pdf.cell(0, 6, hdr, ln=1, align="C")
    pdf.ln(2)

    # --------- MICRO-POLISH HELPERS ----------
    def fmt2(v):
        """format number with commas & 2 decimals"""
        try:
            return f"{float(v):,.2f}"
        except Exception:
            return "0.00"

    def cellR(w, h, txt, **kw):
        """right-aligned numeric cell"""
        pdf.cell(w, h, txt, align="R", **kw)
    # -----------------------------------------

    # headers
    def draw_invoice_header():
        pdf.set_fill_color(230, 240, 255)
        pdf.set_font("Helvetica", "B", 9)
        # widths adjusted to include Firm column
        cols = [
            ("Inv No", 16), ("Date", 20), ("Broker", 30), ("Firm", 30),
            ("Total", 18), ("Batav", 18), ("DR", 14),
            ("Other", 14), ("Adv", 14), ("Net", 18),
        ]
        for text, w in cols:
            pdf.cell(w, 7, text, border=1, align="C", fill=True)
        pdf.ln(7)
        pdf.set_font("Helvetica", "", 9)

    def draw_detail_header():
        pdf.set_fill_color(245, 245, 245)
        pdf.set_font("Helvetica", "B", 8)
        # Adjusted widths to include FrkWt column
        cols = [
            ("Item", 40), ("Bora", 14), ("TBWt", 14),
            ("Qty", 12), ("Rate", 14), ("Amount", 20),
            ("PWt", 14), ("MWt", 14), ("FrkWt", 12), ("DWt", 12), ("Lot", 12),
        ]
        for text, w in cols:
            pdf.cell(w, 6, text, border=1, align="C", fill=True)
        pdf.ln(6)
        pdf.set_font("Helvetica", "", 8)

    current_group = None
    draw_invoice_header()

    for s in sales:
        key = group_key(s)
        if current_group is None or key != current_group:
            # group band
            pdf.set_font("Helvetica", "B", 9)
            if report_type == "date":
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')}"
            elif report_type == "broker":
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')} - {key[1] or 'No Broker'}"
            else:  # firm grouping
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')} - {key[1] or 'No Firm'}"
            pdf.ln(2)
            pdf.set_fill_color(235, 235, 235)
            pdf.cell(0, 6, grp_txt, ln=1, fill=True)
            pdf.set_font("Helvetica", "", 9)
            current_group = key

        # invoice header row: text left, numbers right
        pdf.cell(16, 7, str(s.invno), border=1, align="C")
        pdf.cell(20, 7, s.invdate.strftime("%d-%m-%Y"), border=1, align="C")
        pdf.cell(30, 7, (s.broker.brokername if s.broker else "")[:28], border=1, align="L")
        pdf.cell(30, 7, (s.firm.firmname if getattr(s, "firm", None) else "")[:28], border=1, align="L")
        cellR(18, 7, fmt2(s.totalamt), border=1)
        cellR(18, 7, fmt2(s.batavamt), border=1)
        cellR(14, 7, fmt2(s.dramt), border=1)
        cellR(14, 7, fmt2(s.other), border=1)
        cellR(14, 7, fmt2(s.advance), border=1)
        cellR(18, 7, fmt2(s.netamt), border=1)
        pdf.ln(7)

        # details
        draw_detail_header()
        for d in s.details.all():
            pdf.cell(40, 6, (d.item.item_name or "")[:28], border=1, align="L")
            cellR(14, 6, fmt2(d.bora), border=1)
            cellR(14, 6, fmt2(d.tbwt), border=1)
            cellR(12, 6, fmt2(d.qty), border=1)
            cellR(14, 6, fmt2(d.rate), border=1)
            cellR(20, 6, fmt2(d.amount), border=1)
            cellR(14, 6, fmt2(d.partywt), border=1)
            cellR(14, 6, fmt2(d.millwt), border=1)
            # FrkWt column
            cellR(12, 6, fmt2(getattr(d, "frkwt", 0)), border=1)
            cellR(12, 6, fmt2(d.diffwt), border=1)
            pdf.cell(12, 6, (d.lotno or "")[:8], border=1, align="C")
            pdf.ln(6)

    # overall totals
    overall = sales.aggregate(
        total_totalamt=Sum("totalamt"),
        total_batavamt=Sum("batavamt"),
        total_dramt=Sum("dramt"),
        total_other=Sum("other"),
        total_advance=Sum("advance"),
        total_netamt=Sum("netamt"),
    )
    overall_tbwt = SaleDetails.objects.filter(salemaster__in=sales).aggregate(
        total_tbwt=Sum("tbwt")
    )["total_tbwt"] or 0
    overall_frkwt = SaleDetails.objects.filter(salemaster__in=sales).aggregate(
        total_frkwt=Sum("frkwt")
    )["total_frkwt"] or 0

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "Overall Totals", ln=1)
    pdf.set_font("Helvetica", "", 9)
    lines = [
        f"Total Amt: {fmt2(overall['total_totalamt'] or 0)}",
        f"Batav Amt: {fmt2(overall['total_batavamt'] or 0)}",
        f"DR Amt: {fmt2(overall['total_dramt'] or 0)}",
        f"Other: {fmt2(overall['total_other'] or 0)}",
        f"Advance: {fmt2(overall['total_advance'] or 0)}",
        f"Total TBWt: {fmt2(overall_tbwt)}",
        f"Total FrkWt: {fmt2(overall_frkwt)}",
        f"Net Amt: {fmt2(overall['total_netamt'] or 0)}",
    ]
    for line in lines:
        pdf.cell(0, 6, line, ln=1)

    # finalize (bytes -> HttpResponse)
        # finalize (bytes/str -> HttpResponse) — robust for different fpdf versions
    pdf.alias_nb_pages()
    filename = f"sale_report_{start_date}_{end_date}.pdf"

    out = pdf.output(dest="S")  # may return str, bytes or bytearray
    if isinstance(out, str):
        pdf_bytes = out.encode("latin-1", "replace")
    elif isinstance(out, bytearray):
        pdf_bytes = bytes(out)
    else:
        # already bytes
        pdf_bytes = out

    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


def sale_search_view(request):
    """
    SaleDetails search tailored to your models.
    GET params supported:
      - frkwt (exact)
      - frkwt_min, frkwt_max (optional range)
      - lotno
      - partyname (matches SaleMaster.party.partyname)
      - invno (matches SaleMaster.invno)
    Orders by SaleMaster.invdate desc. Limits to 500 results.
    """
    qs = SaleDetails.objects.select_related('salemaster', 'salemaster__party').all()

    # numeric filters for FrkWt
    frkwt = request.GET.get('frkwt')
    frkwt_min = request.GET.get('frkwt_min')
    frkwt_max = request.GET.get('frkwt_max')

    if frkwt:
        try:
            # cast to float/Decimal compare (Django will handle string numbers too)
            qs = qs.filter(frkwt__exact=frkwt)
        except (ValueError, TypeError):
            pass
    else:
        if frkwt_min:
            try:
                qs = qs.filter(frkwt__gte=frkwt_min)
            except (ValueError, TypeError):
                pass
        if frkwt_max:
            try:
                qs = qs.filter(frkwt__lte=frkwt_max)
            except (ValueError, TypeError):
                pass

    # string-based filters
    lotno = request.GET.get('lotno')
    if lotno:
        qs = qs.filter(lotno__icontains=lotno)

    partyname = request.GET.get('partyname')
    if partyname:
        # your HeadParty field is "partyname"
        qs = qs.filter(salemaster__party__partyname__icontains=partyname)

    invno = request.GET.get('invno')
    if invno:
        qs = qs.filter(salemaster__invno__icontains=invno)

    # final ordering and limit
    sales = qs.order_by('-salemaster__invdate')[:500]

    return render(request, 'brokerapp/sale_search.html', {'sales': sales})


def bardana_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    party_id = request.GET.get("party")
    broker_id = request.GET.get("broker")
    report_type = request.GET.get("report_type", "date")  # date / party / broker

    # Default date = today
    if not start_date:
        start_date = date.today().strftime("%Y-%m-%d")
    if not end_date:
        end_date = date.today().strftime("%Y-%m-%d")

    # Base queryset (ORG SCOPED via salemaster__org)
    details = (
        SaleDetails.objects
        .select_related('salemaster', 'item', 'salemaster__party', 'salemaster__broker')
        .filter(
            salemaster__org=request.current_org,
            salemaster__invdate__gte=parse_date(start_date),
            salemaster__invdate__lte=parse_date(end_date)
        )
        .order_by('salemaster__invdate')
    )

    # Party / Broker filters (within org)
    if party_id and party_id != "all":
        details = details.filter(salemaster__party__pk=party_id)
    if broker_id and broker_id != "all":
        # allow pk or name
        if details.filter(salemaster__broker__pk=broker_id).exists():
            details = details.filter(salemaster__broker__pk=broker_id)
        else:
            details = details.filter(salemaster__broker__brokername=broker_id)

    # Prepare grouped data
    report_data = []

    if report_type == "date":
        group_values = details.values_list('salemaster__invdate', flat=True).distinct()
        for g in group_values:
            group_details = details.filter(salemaster__invdate=g)
            totals = group_details.aggregate(total_bn=Sum('bn'), total_bo=Sum('bo'))
            report_data.append({
                "group": g.strftime("%d-%m-%Y"),
                "items": group_details,
                "total_bn": totals["total_bn"] or 0,
                "total_bo": totals["total_bo"] or 0,
            })

    elif report_type == "party":
        group_values = details.values_list('salemaster__party__partyname', flat=True).distinct()
        for g in group_values:
            group_details = details.filter(salemaster__party__partyname=g)
            totals = group_details.aggregate(total_bn=Sum('bn'), total_bo=Sum('bo'))
            report_data.append({
                "group": g,
                "items": group_details,
                "total_bn": totals["total_bn"] or 0,
                "total_bo": totals["total_bo"] or 0,
            })

    elif report_type == "broker":
        group_values = details.values_list('salemaster__broker__brokername', flat=True).distinct()
        for g in group_values:
            group_details = details.filter(salemaster__broker__brokername=g)
            totals = group_details.aggregate(total_bn=Sum('bn'), total_bo=Sum('bo'))
            report_data.append({
                "group": g or "No Broker",
                "items": group_details,
                "total_bn": totals["total_bn"] or 0,
                "total_bo": totals["total_bo"] or 0,
            })

    # ORG-scoped dropdown lists
    parties = HeadParty.objects.filter(org=request.current_org).order_by("partyname")
    brokers = Broker.objects.filter(org=request.current_org).order_by("brokername")

    context = {
        "report_data": report_data,
        "parties": parties,
        "brokers": brokers,
        "start_date": start_date,
        "end_date": end_date,
        "selected_party": party_id if party_id != "all" else None,
        "selected_broker": broker_id if broker_id != "all" else None,
        "report_type": report_type,
    }
    return render(request, "brokerapp/bardana_report.html", context)


def purchase_form(request, invno=None):
    """
    Render purchase form. If invno provided, load purchase + details (scoped to current org).
    """
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchase = None
    purchase_items_json = "[]"
    today_date = date.today().strftime("%Y-%m-%d")

    if invno:
        # must belong to current org
        purchase = get_object_or_404(PurchaseMaster, invno=invno, org=request.current_org)
        details = PurchaseDetails.objects.filter(purchasemaster=purchase)
        items_data = []
        for d in details:
            items_data.append({
                "item_id": d.item.pk,
                "item_name": d.item.item_name,
                "bora": float(d.bora),
                "bn": float(d.bn),
                "bnwt": float(d.bnwt),
                "bo": float(d.bo),
                "bowt": float(d.bowt),
                "tbwt": float(getattr(d, "tbwt", 0)),          # NEW: TBWt like sale
                "totalbora": float(d.bn * d.bnwt + d.bo * d.bowt),
                "qty": float(d.qty),
                "rate": float(d.rate),
                "amt": float(d.amount),
                "partywt": float(d.partywt),
                "millwt": float(d.millwt),
                "diffwt": float(d.diffwt),
                "frkwt": float(getattr(d, "frkwt", 0)),        # NEW: FrkWt like sale
                "lotno": d.lotno or "",
            })
        purchase_items_json = json.dumps(items_data)

    # next invoice number — per ORG
    next_invno = PurchaseMaster.objects.filter(org=request.current_org).aggregate(Max("invno"))['invno__max']
    next_invno = (next_invno + 1) if next_invno else 1

    # Firms: try to scope to current_org if Firm has org FK, otherwise return all firms
    try:
        firms_qs = Firm.objects.filter(org=request.current_org).order_by('firmname')
    except Exception:
        firms_qs = Firm.objects.all().order_by('firmname')

    context = {
        "purchase": purchase,
        "purchase_items_json": purchase_items_json,
        "next_invno": next_invno,
        "today_date": today_date,
        # only current org choices
        "items": HeadItem.objects.filter(org=request.current_org).order_by('item_name'),
        "parties": HeadParty.objects.filter(org=request.current_org).order_by('partyname'),
        "brokers": Broker.objects.filter(org=request.current_org).order_by('brokername'),
        "firms": firms_qs,
    }
    return render(request, "brokerapp/purchase.html", context)



@transaction.atomic
def save_purchase(request):
    """
    Save a new PurchaseMaster and its PurchaseDetails — scoped to current org.
    Also sends invoice email with attached PDF (like Sale).
    """
    assert getattr(request, "current_org", None) is not None, "current_org missing"

    if request.method != "POST":
        return redirect("purchase_form_new")

    try:
        # -------- Header fields --------
        invdate_str = request.POST.get("invdate")
        invdate = (
            datetime.strptime(invdate_str, "%Y-%m-%d").date()
            if invdate_str else date.today()
        )
        awakno = request.POST.get("awakno", "").strip()
        extra = request.POST.get("extra", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        firm_pk = request.POST.get("firm")
        vehicleno = request.POST.get("vehicleno", "").strip()

        # -------- Items JSON --------
        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("purchase_form_new")

        # -------- Totals (same style as Sale) --------
        total_amt = Decimal("0")
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal("100")).quantize(Decimal("0.01"))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal("100")).quantize(Decimal("0.01"))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))

        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal("0.01"))
        netamt = (total - advance).quantize(Decimal("0.01"))

        # -------- Resolve FKs within same org --------
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        firm = None
        if firm_pk:
            firm_fields = [f.name for f in Firm._meta.get_fields()]
            if "org" in firm_fields:
                firm = get_object_or_404(Firm, pk=firm_pk, org=request.current_org)
            else:
                firm = get_object_or_404(Firm, pk=firm_pk)

        # -------- Create PurchaseMaster --------
        purchase = PurchaseMaster.objects.create(
            org=request.current_org,
            created_by=request.user,
            invdate=invdate,
            awakno=awakno,
            party=party,
            broker=broker,
            firm=firm,
            vehicleno=vehicleno,
            extra=extra,
            totalamt=total_amt.quantize(Decimal("0.01")),
            batavpercent=batavpercent,
            batavamt=batavamt,
            dr=dr,
            dramt=dramt,
            qi=qi,
            other=other,
            total=total,
            advance=advance,
            netamt=netamt,
            remark=request.POST.get("remark", "").strip(),
        )

        # -------- Create PurchaseDetails (per item) --------
        for it in items:
            item_id = it.get("item_id")
            item_obj = get_object_or_404(HeadItem, pk=item_id, org=request.current_org)

            PurchaseDetails.objects.create(
                purchasemaster=purchase,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                tbwt=to_decimal(it.get("tbwt", 0)),      # TBWt field in model
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                frkwt=to_decimal(it.get("frkwt", 0)),    # FrkWt field in model
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        # -------------------------
        #  EMAIL + PROFESSIONAL PDF INVOICE (like Sale)
        # -------------------------
        send_email = request.POST.get("send_email") == "on"

        if send_email and party.email:
            if FPDF is None:
                messages.warning(
                    request,
                    "Invoice email ke liye 'fpdf' package chahiye. Install: pip install fpdf"
                )
            else:
                def safe_text(val, maxlen=None):
                    s = "" if val is None else str(val)
                    s = s.replace("—", "-").replace("–", "-")
                    s = "".join(ch if ord(ch) < 128 else "?" for ch in s)
                    return s[:maxlen] if maxlen else s

                try:
                    # ---------- 1) PDF ----------
                    pdf = FPDF("P", "mm", "A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.set_margins(15, 15, 15)
                    pdf.add_page()

                    # Top org name
                    org_name = getattr(purchase.org, "name", str(purchase.org))
                    pdf.set_font("Helvetica", "B", 16)
                    pdf.cell(0, 8, safe_text(org_name, 60), ln=1)

                    pdf.set_font("Helvetica", "", 10)
                    pdf.cell(0, 6, safe_text("Purchase Invoice", 40), ln=1)
                    pdf.ln(2)

                    # Invoice meta
                    pdf.set_font("Helvetica", "", 10)
                    pdf.cell(28, 6, "Invoice No:", 0, 0)
                    pdf.cell(60, 6, safe_text(purchase.invno), 0, 0)
                    pdf.cell(20, 6, "Date:", 0, 0)
                    pdf.cell(0, 6, safe_text(purchase.invdate.strftime("%d-%m-%Y")), 0, 1)

                    pdf.cell(28, 6, "Party:", 0, 0)
                    pdf.cell(60, 6, safe_text(purchase.party.partyname, 40), 0, 0)
                    pdf.cell(20, 6, "Vehicle:", 0, 0)
                    pdf.cell(0, 6, safe_text(purchase.vehicleno or "-", 30), 0, 1)

                    pdf.cell(28, 6, "Firm:", 0, 0)
                    firm_label = getattr(purchase.firm, "firmname", "") if purchase.firm else ""
                    pdf.cell(60, 6, safe_text(firm_label, 40), 0, 0)
                    pdf.cell(20, 6, "Broker:", 0, 0)
                    pdf.cell(0, 6, safe_text(purchase.broker.brokername, 40), 0, 1)

                    pdf.ln(4)

                    # Items table
                    headers = ["Item", "Bora", "TBWt", "Qty", "FrkWt", "Rate", "Amount", "LotNo"]
                    widths  = [45,   15,    18,    15,    18,     22,    25,      22]

                    pdf.set_font("Helvetica", "B", 9)
                    for i, h in enumerate(headers):
                        pdf.cell(widths[i], 7, safe_text(h, 20), border=1, align="C")
                    pdf.ln(7)

                    pdf.set_font("Helvetica", "", 9)
                    details_qs = PurchaseDetails.objects.filter(purchasemaster=purchase).select_related("item")

                    for d in details_qs:
                        row = [
                            safe_text(d.item.item_name, 30),
                            safe_text(f"{d.bora:.0f}", 5),
                            safe_text(f"{d.tbwt:.2f}", 10),
                            safe_text(f"{d.qty:.2f}", 10),
                            safe_text(f"{d.frkwt:.2f}", 10),
                            safe_text(f"{d.rate:.2f}", 10),
                            safe_text(f"{d.amount:.2f}", 12),
                            safe_text(d.lotno or "", 12),
                        ]
                        for i, v in enumerate(row):
                            align = "R" if i in (1, 2, 3, 4, 5, 6) else "L"
                            pdf.cell(widths[i], 7, v, border=1, align=align)
                        pdf.ln(7)

                    pdf.ln(3)

                    # Totals on right side
                    pdf.set_font("Helvetica", "", 9)
                    right_x = pdf.w - 15 - 60   # 60mm wide totals box
                    pdf.set_xy(right_x, pdf.get_y())

                    lines = [
                        ("Total Amount",    purchase.totalamt),
                        (f"Batav ({purchase.batavpercent:.2f}%)", purchase.batavamt),
                        (f"Dr ({purchase.dr:.2f}%)", purchase.dramt),
                        ("QI + Other",      purchase.qi + purchase.other),
                        ("Advance",         purchase.advance),
                        ("Net Amount",      purchase.netamt),
                    ]
                    for label, val in lines:
                        pdf.cell(35, 6, safe_text(label, 30), border=0, align="R")
                        pdf.cell(25, 6, f"{val:.2f}", border=0, align="R")
                        pdf.ln(6)

                    # Footer
                    pdf.ln(4)
                    pdf.set_x(15)
                    pdf.set_font("Helvetica", "I", 8)
                    pdf.cell(0, 5, "This is a system generated invoice.", 0, 1)

                    buf = io.BytesIO()
                    pdf.output(buf)
                    buf.seek(0)
                    pdf_bytes = buf.read()

                    safe_party = "".join(
                        ch if ord(ch) < 128 else "?" for ch in party.partyname
                    )[:40]

                    # ---------- 2) Email with attached PDF ----------
                    subject = f"Purchase Invoice #{purchase.invno}"
                    body = (
                        f"Dear {party.partyname},\n\n"
                        "Please find your purchase invoice attached as PDF.\n\n"
                        f"Total Amount: {purchase.totalamt:.2f}\n"
                        f"Net Amount  : {purchase.netamt:.2f}\n\n"
                        "Thank you."
                    )

                    email = EmailMessage(
                        subject,
                        body,
                        settings.DEFAULT_FROM_EMAIL,
                        [party.email],
                    )
                    email.attach(
                        f"purchase_{purchase.invno}_{safe_party}.pdf",
                        pdf_bytes,
                        "application/pdf",
                    )
                    email.send(fail_silently=False)

                except Exception as e:
                    messages.warning(request, f"Purchase saved but invoice email not sent: {e}")

        messages.success(request, "Purchase entry saved successfully!")
        return redirect("purchasedata")

    except Exception as e:
        messages.error(request, f"Error saving purchase: {e}")
        return redirect("purchase_form_new")


@transaction.atomic
def update_purchase(request, invno):
    """
    Update an existing PurchaseMaster and its PurchaseDetails — with email + PDF.
    """
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchase = get_object_or_404(PurchaseMaster, invno=invno, org=request.current_org)

    if request.method != "POST":
        return redirect("purchase_form_update", invno=invno)

    try:
        # ---------- HEADER ----------
        invdate_str = request.POST.get("invdate")
        invdate = (
            datetime.strptime(invdate_str, "%Y-%m-%d").date()
            if invdate_str else purchase.invdate
        )
        awakno = request.POST.get("awakno", "").strip()
        party_pk = request.POST.get("party")
        broker_pk = request.POST.get("broker")
        vehicleno = request.POST.get("vehicleno", "").strip()
        extra = request.POST.get("extra", "").strip()

        # ---------- ITEMS ----------
        items_json = request.POST.get("items_json") or "[]"
        items = json.loads(items_json)
        if not items:
            messages.error(request, "Add at least one item before saving.")
            return redirect("purchase_form_update", invno=invno)

        total_amt = Decimal("0")
        for it in items:
            total_amt += to_decimal(it.get("amt", 0))

        batavpercent = to_decimal(request.POST.get("batavpercent", 0))
        batavamt = (total_amt * batavpercent / Decimal("100")).quantize(Decimal("0.01"))

        dr = to_decimal(request.POST.get("dr", 0))
        dramt = (total_amt * dr / Decimal("100")).quantize(Decimal("0.01"))

        qi = to_decimal(request.POST.get("qi", 0))
        other = to_decimal(request.POST.get("other", 0))
        advance = to_decimal(request.POST.get("advance", 0))

        total = (total_amt - batavamt - dramt - qi - other).quantize(Decimal("0.01"))
        netamt = (total - advance).quantize(Decimal("0.01"))

        # ---------- FOREIGN KEYS ----------
        party = get_object_or_404(HeadParty, pk=party_pk, org=request.current_org)
        broker = get_object_or_404(Broker, pk=broker_pk, org=request.current_org)

        firm = None
        firm_pk = request.POST.get("firm")
        if firm_pk:
            firm_fields = [f.name for f in Firm._meta.get_fields()]
            if "org" in firm_fields:
                firm = get_object_or_404(Firm, pk=firm_pk, org=request.current_org)
            else:
                firm = get_object_or_404(Firm, pk=firm_pk)

        # ---------- UPDATE HEADER ----------
        purchase.invdate = invdate
        purchase.awakno = awakno
        purchase.party = party
        purchase.broker = broker
        purchase.vehicleno = vehicleno
        purchase.extra = extra
        purchase.totalamt = total_amt.quantize(Decimal("0.01"))
        purchase.batavpercent = batavpercent
        purchase.batavamt = batavamt
        purchase.dr = dr
        purchase.dramt = dramt
        purchase.qi = qi
        purchase.other = other
        purchase.advance = advance
        purchase.total = total
        purchase.netamt = netamt
        purchase.remark = request.POST.get("remark", "").strip()
        purchase.firm = firm
        purchase.save()

        # ---------- REPLACE DETAILS ----------
        PurchaseDetails.objects.filter(purchasemaster=purchase).delete()

        for it in items:
            item_obj = get_object_or_404(
                HeadItem, pk=it.get("item_id"), org=request.current_org
            )
            PurchaseDetails.objects.create(
                purchasemaster=purchase,
                item=item_obj,
                bora=to_decimal(it.get("bora", 0)),
                bn=to_decimal(it.get("bn", 0)),
                bnwt=to_decimal(it.get("bnwt", 0)),
                bo=to_decimal(it.get("bo", 0)),
                bowt=to_decimal(it.get("bowt", 0)),
                tbwt=to_decimal(it.get("tbwt", 0)),
                qty=to_decimal(it.get("qty", 0)),
                rate=to_decimal(it.get("rate", 0)),
                amount=to_decimal(it.get("amt", 0)),
                partywt=to_decimal(it.get("partywt", 0)),
                millwt=to_decimal(it.get("millwt", 0)),
                frkwt=to_decimal(it.get("frkwt", 0)),
                diffwt=to_decimal(it.get("diffwt", 0)),
                lotno=it.get("lotno", "").strip(),
            )

        # ==========================================================
        #  EMAIL + PDF AFTER UPDATE (Same as save_purchase)
        # ==========================================================
        send_email = request.POST.get("send_email") == "on"

        if send_email and party.email:
            if FPDF is None:
                messages.warning(
                    request, "Invoice email ke liye FPDF install kare: pip install fpdf"
                )
            else:
                def safe_text(v, maxlen=None):
                    s = "" if v is None else str(v)
                    s = "".join(ch if ord(ch) < 128 else "?" for ch in s)
                    return s[:maxlen] if maxlen else s

                try:
                    # --- PDF ---
                    pdf = FPDF("P", "mm", "A4")
                    pdf.set_auto_page_break(auto=True, margin=15)
                    pdf.set_margins(15, 15, 15)
                    pdf.add_page()

                    # Header
                    pdf.set_font("Helvetica", "B", 16)
                    pdf.cell(0, 8, safe_text(purchase.org.name), ln=1)

                    pdf.set_font("Helvetica", "", 10)
                    pdf.cell(0, 6, "Purchase Invoice (Updated)", ln=1)

                    pdf.ln(2)

                    # Meta
                    pdf.cell(30, 6, "Invoice No:", 0, 0)
                    pdf.cell(50, 6, safe_text(purchase.invno), 0, 0)
                    pdf.cell(20, 6, "Date:", 0, 0)
                    pdf.cell(0, 6, purchase.invdate.strftime("%d-%m-%Y"), 0, 1)

                    pdf.cell(30, 6, "Party:", 0, 0)
                    pdf.cell(50, 6, safe_text(party.partyname), 0, 0)
                    pdf.cell(20, 6, "Vehicle:", 0, 0)
                    pdf.cell(0, 6, safe_text(purchase.vehicleno or "-"), 0, 1)

                    pdf.cell(30, 6, "Firm:", 0, 0)
                    pdf.cell(50, 6, safe_text(purchase.firm.firmname if purchase.firm else ""), 0, 0)
                    pdf.cell(20, 6, "Broker:", 0, 0)
                    pdf.cell(0, 6, safe_text(broker.brokername), 0, 1)

                    pdf.ln(4)

                    # Items table
                    headers = ["Item", "Bora", "TBWt", "Qty", "FrkWt", "Rate", "Amt", "Lot"]
                    widths  = [45, 15, 18, 15, 18, 22, 25, 22]

                    pdf.set_font("Helvetica", "B", 9)
                    for i, h in enumerate(headers):
                        pdf.cell(widths[i], 7, h, border=1, align="C")
                    pdf.ln(7)

                    pdf.set_font("Helvetica", "", 9)
                    for d in purchase.details.all():
                        row = [
                            safe_text(d.item.item_name, 30),
                            f"{d.bora:.0f}",
                            f"{d.tbwt:.2f}",
                            f"{d.qty:.2f}",
                            f"{d.frkwt:.2f}",
                            f"{d.rate:.2f}",
                            f"{d.amount:.2f}",
                            safe_text(d.lotno),
                        ]
                        for i, v in enumerate(row):
                            align = "R" if i > 0 else "L"
                            pdf.cell(widths[i], 7, v, border=1, align=align)
                        pdf.ln(7)

                    # Totals
                    pdf.ln(4)
                    pdf.set_x(pdf.w - 75)
                    pairs = [
                        ("Total", purchase.totalamt),
                        (f"Batav {purchase.batavpercent}%", purchase.batavamt),
                        (f"DR {purchase.dr}%", purchase.dramt),
                        ("QI + Other", purchase.qi + purchase.other),
                        ("Advance", purchase.advance),
                        ("Net Amount", purchase.netamt),
                    ]
                    for label, val in pairs:
                        pdf.cell(35, 6, label, 0, 0, "R")
                        pdf.cell(25, 6, f"{val:.2f}", 0, 1, "R")

                    # Convert to bytes
                    buf = io.BytesIO()
                    pdf.output(buf)
                    buf.seek(0)
                    pdf_bytes = buf.read()

                    # Email
                    subject = f"Updated Purchase Invoice #{purchase.invno}"
                    body = (
                        f"Dear {party.partyname},\n\n"
                        "Your updated purchase invoice is attached.\n\n"
                        f"Net Amount: {purchase.netamt:.2f}\n\n"
                        "Thank you."
                    )

                    email = EmailMessage(
                        subject,
                        body,
                        settings.DEFAULT_FROM_EMAIL,
                        [party.email],
                    )
                    email.attach(
                        f"purchase_{purchase.invno}.pdf",
                        pdf_bytes,
                        "application/pdf",
                    )
                    email.send()

                except Exception as e:
                    messages.warning(request, f"Invoice updated but email not sent: {e}")

        messages.success(request, "Purchase entry updated successfully!")
        return redirect("purchasedata")

    except Exception as e:
        messages.error(request, f"Error updating purchase: {e}")
        return redirect("purchase_form_update", invno=invno)

def purchase_data_view(request):
    """List of purchases (scoped to current org)."""
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchases = PurchaseMaster.objects.filter(org=request.current_org).order_by("-invno")
    return render(request, "brokerapp/purchasedata.html", {
        "purchases": purchases,
        "today_date": date.today(),
    })




def delete_purchase(request, invno):
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    purchase = get_object_or_404(PurchaseMaster, invno=invno, org=request.current_org)
    purchase.delete()
    messages.success(request, "Purchase entry deleted successfully!")
    return redirect("purchasedata")

def purchase_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    broker_id = request.GET.get("broker")
    firm_id = request.GET.get("firm")         # optional firm filter (pk or name)
    report_type = request.GET.get("report_type", "date")

    # Default date = today
    if not start_date:
        start_date = date.today().strftime("%Y-%m-%d")
    if not end_date:
        end_date = date.today().strftime("%Y-%m-%d")

    # Base queryset (ORG SCOPED) + prefetch details (with item) for the template
    purchases = (
        PurchaseMaster.objects
        .filter(org=request.current_org)
        .select_related("broker", "firm")   # firm selected for template access
        .prefetch_related(Prefetch("details", queryset=PurchaseDetails.objects.select_related("item")))
    )

    if start_date:
        purchases = purchases.filter(invdate__gte=parse_date(start_date))
    if end_date:
        purchases = purchases.filter(invdate__lte=parse_date(end_date))

    # Broker filter: allow pk or name, but scoped to org
    if broker_id and broker_id != "all":
        # try as primary key
        if purchases.filter(broker__pk=broker_id).exists():
            purchases = purchases.filter(broker__pk=broker_id)
        else:
            purchases = purchases.filter(broker__brokername=broker_id)

    # Firm filter: allow pk or name, scoped to org (only apply when provided and not "all")
    if firm_id and firm_id != "all":
        # try pk (firm FK)
        if purchases.filter(firm__pk=firm_id).exists():
            purchases = purchases.filter(firm__pk=firm_id)
        else:
            purchases = purchases.filter(firm__firmname=firm_id)

    purchases = purchases.order_by("invdate")

    report_data = []

    # --- Grouping & Aggregates ---
    if report_type == "date":
        grouped = purchases.values("invdate").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate")

        for g in grouped:
            group_purchases = purchases.filter(invdate=g["invdate"])
            # TBWt & FrkWt sum for this group (sum over details)
            tbwt_sum = PurchaseDetails.objects.filter(purchasemaster__in=group_purchases).aggregate(
                total_tbwt=Sum("tbwt")
            )["total_tbwt"] or 0
            frkwt_sum = PurchaseDetails.objects.filter(purchasemaster__in=group_purchases).aggregate(
                total_frkwt=Sum("frkwt")
            )["total_frkwt"] or 0

            g["total_tbwt"] = tbwt_sum
            g["total_frkwt"] = frkwt_sum

            report_data.append({
                "group": g["invdate"],
                "items": group_purchases,
                "totals": g
            })

    elif report_type == "broker":
        grouped = purchases.values("invdate", "broker__brokername").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate", "broker__brokername")

        for g in grouped:
            group_purchases = purchases.filter(
                invdate=g["invdate"],
                broker__brokername=g["broker__brokername"]
            )
            tbwt_sum = PurchaseDetails.objects.filter(purchasemaster__in=group_purchases).aggregate(
                total_tbwt=Sum("tbwt")
            )["total_tbwt"] or 0
            frkwt_sum = PurchaseDetails.objects.filter(purchasemaster__in=group_purchases).aggregate(
                total_frkwt=Sum("frkwt")
            )["total_frkwt"] or 0

            g["total_tbwt"] = tbwt_sum
            g["total_frkwt"] = frkwt_sum

            report_data.append({
                "group": f"{g['invdate']} - {g['broker__brokername'] or 'No Broker'}",
                "items": group_purchases,
                "totals": g
            })

    elif report_type == "firm":
        # Group by invdate + firm (firm may be null)
        grouped = purchases.values("invdate", "firm__firmname").annotate(
            total_totalamt=Sum("totalamt"),
            total_batavamt=Sum("batavamt"),
            total_dramt=Sum("dramt"),
            total_other=Sum("other"),
            total_total=Sum("total"),
            total_advance=Sum("advance"),
            total_netamt=Sum("netamt"),
        ).order_by("invdate", "firm__firmname")

        for g in grouped:
            firm_name = g.get("firm__firmname")  # may be None
            if firm_name is None:
                # purchases with NULL firm for that date
                group_purchases = purchases.filter(invdate=g["invdate"], firm__isnull=True)
                label_firm = "No Firm"
            else:
                group_purchases = purchases.filter(invdate=g["invdate"], firm__firmname=firm_name)
                label_firm = firm_name

            tbwt_sum = PurchaseDetails.objects.filter(purchasemaster__in=group_purchases).aggregate(
                total_tbwt=Sum("tbwt")
            )["total_tbwt"] or 0
            frkwt_sum = PurchaseDetails.objects.filter(purchasemaster__in=group_purchases).aggregate(
                total_frkwt=Sum("frkwt")
            )["total_frkwt"] or 0

            g["total_tbwt"] = tbwt_sum
            g["total_frkwt"] = frkwt_sum

            report_data.append({
                "group": f"{g['invdate']} - {label_firm}",
                "items": group_purchases,
                "totals": g
            })

    # Overall Totals (header-level) + TBWt + FrkWt across all details in the filtered set
    overall_totals = purchases.aggregate(
        total_totalamt=Sum("totalamt"),
        total_batavamt=Sum("batavamt"),
        total_dramt=Sum("dramt"),
        total_other=Sum("other"),
        total_total=Sum("total"),
        total_advance=Sum("advance"),
        total_netamt=Sum("netamt"),
    )
    overall_tbwt = PurchaseDetails.objects.filter(purchasemaster__in=purchases).aggregate(
        total_tbwt=Sum("tbwt")
    )["total_tbwt"] or 0
    overall_frkwt = PurchaseDetails.objects.filter(purchasemaster__in=purchases).aggregate(
        total_frkwt=Sum("frkwt")
    )["total_frkwt"] or 0

    overall_totals["total_tbwt"] = overall_tbwt
    overall_totals["total_frkwt"] = overall_frkwt

    # Dropdowns also ORG SCOPED
    brokers = Broker.objects.filter(org=request.current_org).order_by("brokername")

    try:
        firm_field_names = [f.name for f in Firm._meta.get_fields()]
        if 'org' in firm_field_names:
            firms = Firm.objects.filter(org=request.current_org).order_by('firmname')
        else:
            firms = Firm.objects.all().order_by('firmname')
    except Exception:
        firms = Firm.objects.all().order_by('firmname')

    context = {
        "report_data": report_data,
        "overall_totals": overall_totals,
        "start_date": start_date,
        "end_date": end_date,
        "brokers": brokers,
        "firms": firms,
        "selected_broker": broker_id if broker_id != "all" else None,
        "selected_firm": firm_id if firm_id != "all" else None,
        "report_type": report_type,
    }
    return render(request, "brokerapp/purchase_report.html", context)

# ===================== PDF (FPDF) =====================
def purchase_report_pdf(request):
    """
    Generate Purchase Report PDF (FPDF) using current filters.
    Includes per-invoice detail rows with TBWt and FrkWt.
    Numbers are right-aligned with thousand separators.
    """
    # --- build same filtered queryset as HTML purchase report ---
    start_date = request.GET.get("start_date") or date.today().strftime("%Y-%m-%d")
    end_date = request.GET.get("end_date") or date.today().strftime("%Y-%m-%d")
    broker_id = request.GET.get("broker")
    firm_id = request.GET.get("firm")          # firm filter
    report_type = request.GET.get("report_type", "date")

    purchases = (
        PurchaseMaster.objects
        .filter(org=request.current_org)
        .select_related("broker", "firm")   # firm selected for template/pdf access
        .prefetch_related(Prefetch("details", queryset=PurchaseDetails.objects.select_related("item")))
    )
    if start_date:
        purchases = purchases.filter(invdate__gte=parse_date(start_date))
    if end_date:
        purchases = purchases.filter(invdate__lte=parse_date(end_date))

    if broker_id and broker_id != "all":
        if purchases.filter(broker__pk=broker_id).exists():
            purchases = purchases.filter(broker__pk=broker_id)
        else:
            purchases = purchases.filter(broker__brokername=broker_id)

    # APPLY FIRM FILTER (same behavior as in HTML view)
    if firm_id and firm_id != "all":
        if purchases.filter(firm__pk=firm_id).exists():
            purchases = purchases.filter(firm__pk=firm_id)
        else:
            purchases = purchases.filter(firm__firmname=firm_id)

    purchases = purchases.order_by("invdate", "invno")

    # group-key helpers (just for headings)
    if report_type == "date":
        def group_key(p): return (p.invdate,)
    elif report_type == "broker":
        def group_key(p): return (p.invdate, p.broker.brokername if p.broker else "")
    else:  # report_type == "firm"
        def group_key(p): return (p.invdate, p.firm.firmname if p.firm else "")

    # --- FPDF setup ---
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Purchase Report", ln=1, align="C")
    pdf.set_font("Helvetica", "", 9)
    hdr = f"From {start_date} To {end_date} | Generated: {timezone.now().strftime('%d-%m-%Y %I:%M %p')}"
    pdf.cell(0, 6, hdr, ln=1, align="C")
    pdf.ln(2)

    # --------- MICRO-POLISH HELPERS ----------
    def fmt2(v):
        """format number with commas & 2 decimals"""
        try:
            return f"{float(v):,.2f}"
        except Exception:
            return "0.00"

    def cellR(w, h, txt, **kw):
        """right-aligned numeric cell"""
        pdf.cell(w, h, txt, align="R", **kw)
    # -----------------------------------------

    # headers
    def draw_invoice_header():
        pdf.set_fill_color(230, 240, 255)
        pdf.set_font("Helvetica", "B", 9)
        # widths adjusted to include Firm column
        cols = [
            ("Inv No", 16), ("Date", 20), ("Broker", 30), ("Firm", 30),
            ("Total", 18), ("Batav", 18), ("DR", 14),
            ("Other", 14), ("Adv", 14), ("Net", 18),
        ]
        for text, w in cols:
            pdf.cell(w, 7, text, border=1, align="C", fill=True)
        pdf.ln(7)
        pdf.set_font("Helvetica", "", 9)

    def draw_detail_header():
        pdf.set_fill_color(245, 245, 245)
        pdf.set_font("Helvetica", "B", 8)
        # Adjusted widths to include FrkWt column
        cols = [
            ("Item", 40), ("Bora", 14), ("TBWt", 14),
            ("Qty", 12), ("Rate", 14), ("Amount", 20),
            ("PWt", 14), ("MWt", 14), ("FrkWt", 12), ("DWt", 12), ("Lot", 12),
        ]
        for text, w in cols:
            pdf.cell(w, 6, text, border=1, align="C", fill=True)
        pdf.ln(6)
        pdf.set_font("Helvetica", "", 8)

    current_group = None
    draw_invoice_header()

    for p in purchases:
        key = group_key(p)
        if current_group is None or key != current_group:
            # group band
            pdf.set_font("Helvetica", "B", 9)
            if report_type == "date":
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')}"
            elif report_type == "broker":
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')} - {key[1] or 'No Broker'}"
            else:  # firm grouping
                grp_txt = f"Group: {key[0].strftime('%d-%m-%Y')} - {key[1] or 'No Firm'}"
            pdf.ln(2)
            pdf.set_fill_color(235, 235, 235)
            pdf.cell(0, 6, grp_txt, ln=1, fill=True)
            pdf.set_font("Helvetica", "", 9)
            current_group = key

        # invoice header row
        pdf.cell(16, 7, str(p.invno), border=1, align="C")
        pdf.cell(20, 7, p.invdate.strftime("%d-%m-%Y"), border=1, align="C")
        pdf.cell(30, 7, (p.broker.brokername if p.broker else "")[:28], border=1, align="L")
        pdf.cell(30, 7, (p.firm.firmname if getattr(p, "firm", None) else "")[:28], border=1, align="L")
        cellR(18, 7, fmt2(p.totalamt), border=1)
        cellR(18, 7, fmt2(p.batavamt), border=1)
        cellR(14, 7, fmt2(p.dramt), border=1)
        cellR(14, 7, fmt2(p.other), border=1)
        cellR(14, 7, fmt2(p.advance), border=1)
        cellR(18, 7, fmt2(p.netamt), border=1)
        pdf.ln(7)

        # details
        draw_detail_header()
        for d in p.details.all():
            pdf.cell(40, 6, (d.item.item_name or "")[:28], border=1, align="L")
            cellR(14, 6, fmt2(d.bora), border=1)
            cellR(14, 6, fmt2(d.tbwt), border=1)
            cellR(12, 6, fmt2(d.qty), border=1)
            cellR(14, 6, fmt2(d.rate), border=1)
            cellR(20, 6, fmt2(d.amount), border=1)
            cellR(14, 6, fmt2(d.partywt), border=1)
            cellR(14, 6, fmt2(d.millwt), border=1)
            cellR(12, 6, fmt2(getattr(d, "frkwt", 0)), border=1)  # FrkWt
            cellR(12, 6, fmt2(d.diffwt), border=1)
            pdf.cell(12, 6, (d.lotno or "")[:8], border=1, align="C")
            pdf.ln(6)

    # overall totals
    overall = purchases.aggregate(
        total_totalamt=Sum("totalamt"),
        total_batavamt=Sum("batavamt"),
        total_dramt=Sum("dramt"),
        total_other=Sum("other"),
        total_advance=Sum("advance"),
        total_netamt=Sum("netamt"),
    )
    overall_tbwt = PurchaseDetails.objects.filter(purchasemaster__in=purchases).aggregate(
        total_tbwt=Sum("tbwt")
    )["total_tbwt"] or 0
    overall_frkwt = PurchaseDetails.objects.filter(purchasemaster__in=purchases).aggregate(
        total_frkwt=Sum("frkwt")
    )["total_frkwt"] or 0

    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, "Overall Totals", ln=1)
    pdf.set_font("Helvetica", "", 9)
    lines = [
        f"Total Amt: {fmt2(overall['total_totalamt'] or 0)}",
        f"Batav Amt: {fmt2(overall['total_batavamt'] or 0)}",
        f"DR Amt: {fmt2(overall['total_dramt'] or 0)}",
        f"Other: {fmt2(overall['total_other'] or 0)}",
        f"Advance: {fmt2(overall['total_advance'] or 0)}",
        f"Total TBWt: {fmt2(overall_tbwt)}",
        f"Total FrkWt: {fmt2(overall_frkwt)}",
        f"Net Amt: {fmt2(overall['total_netamt'] or 0)}",
    ]
    for line in lines:
        pdf.cell(0, 6, line, ln=1)

    # finalize (bytes/str/bytearray -> HttpResponse) — robust for different fpdf versions
    pdf.alias_nb_pages()
    filename = f"purchase_report_{start_date}_{end_date}.pdf"

    out = pdf.output(dest="S")  # may return str, bytes or bytearray
    if isinstance(out, str):
        pdf_bytes = out.encode("latin-1", "replace")
    elif isinstance(out, bytearray):
        pdf_bytes = bytes(out)
    else:
        pdf_bytes = out  # already bytes

    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp



def party_view(request, pk=None):
    # Only fetch inside current org
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    instance = None
    if pk:
        instance = get_object_or_404(HeadParty, pk=pk, org=request.current_org)

    if request.method == 'POST':
        # ⬇️ current_org पास करें
        form = PartyForm(request.POST, instance=instance, current_org=request.current_org)

        if form.is_valid():
            obj = form.save(commit=False)

            # Ensure party always belongs to selected org
            obj.org = request.current_org
            obj.save()

            if pk:
                messages.success(request, '✅ Party updated successfully!')
            else:
                messages.success(request, '✅ Party added successfully!')

            # ✅ Redirect logic simplified for Sale return
            next_page = request.GET.get('next')
            if next_page == 'sale':
                # user came from Sale page → go back to Sale after saving
                return redirect(reverse('sale_form_new'))
            elif next_page == 'purchase':
                return redirect(reverse('purchase_form_new'))
            elif next_page == 'daily':
                return redirect(reverse('daily_page'))

            # Default: stay on Party list (existing behavior)
            return redirect(f"{reverse('party')}?created_id={obj.pk}&created_name={quote(obj.partyname)}")

    else:
        # ⬇️ current_org पास करें
        form = PartyForm(instance=instance, current_org=request.current_org)

    # Show only current org parties
    parties = HeadParty.objects.filter(org=request.current_org)

    return render(request, 'brokerapp/party.html', {
        'form': form,
        'parties': parties,
        'editing': pk is not None,
        'editing_id': pk
    })



def party_delete(request, pk):
    """Safely delete a Party — show message if linked to transactions."""
    party = get_object_or_404(HeadParty, pk=pk)
    party_name = party.partyname  # ✅ Save name before deleting
    try:
        party.delete()
        messages.success(request, f"✅ Party '{party_name}' deleted successfully!")
    except ProtectedError:
        messages.error(
            request,
            f"⚠️ Cannot delete '{party_name}' — it is linked to existing Jama or other entries."
        )
    return redirect('party')


def broker_view(request, pk=None):
    # Only fetch inside current org (same as party_view)
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    instance = None
    if pk:
        instance = get_object_or_404(Broker, pk=pk, org=request.current_org)

    if request.method == 'POST':
        # ⬇️ pass current_org into form
        form = BrokerForm(request.POST, instance=instance, current_org=request.current_org)

        if form.is_valid():
            obj = form.save(commit=False)

            # Ensure broker always belongs to selected org
            obj.org = request.current_org
            obj.save()

            if pk:
                messages.success(request, '✅ Broker updated successfully!')
            else:
                messages.success(request, '✅ Broker added successfully!')

            # ✅ Redirect logic simplified for Sale return
            next_page = request.GET.get('next')
            if next_page == 'sale':
                # user came from Sale page → go back to Sale after saving
                return redirect(reverse('sale_form_new'))
            elif next_page == 'purchase':
                return redirect(reverse('purchase_form_new'))
            elif next_page == 'daily':
                return redirect(reverse('daily_page'))

            # Default: stay on Broker list (existing behavior)
            return redirect(f"{reverse('broker')}?created_id={obj.pk}&created_name={quote(obj.brokername)}")

    else:
        # ⬇️ pass current_org into form
        form = BrokerForm(instance=instance, current_org=request.current_org)

    # Show only current org brokers
    brokers = Broker.objects.filter(org=request.current_org)

    return render(request, 'brokerapp/broker.html', {
        'form': form,
        'brokers': brokers,
        'editing': pk is not None,
        'editing_id': pk
    })
# Delete Broker
def broker_delete(request, pk):
    """Safely delete a Broker — show warning if linked to transactions."""
    broker = get_object_or_404(Broker, pk=pk)
    broker_name = broker.brokername  # ✅ store name before delete

    try:
        broker.delete()
        messages.success(request, f"✅ Broker '{broker_name}' deleted successfully!")
    except ProtectedError:
        messages.error(
            request,
            f"⚠️ Cannot delete '{broker_name}' — it is linked to existing Jama or Naame entries."
        )

    return redirect('broker')

def firm_view(request, pk=None):
    """
    Manage firms. Supports `?next=sale|purchase|daily` to return to calling form after save.
    """
    form = FirmForm()
    firms = Firm.objects.all().order_by("firmname")

    if request.method == 'POST':
        form = FirmForm(request.POST)

        if request.POST.get('action') == 'save':
            if form.is_valid():
                # keep behavior similar to other masters: allow attaching org if present
                obj = form.save(commit=False)
                if hasattr(obj, 'org') and getattr(request, 'current_org', None) is not None:
                    obj.org = request.current_org
                obj.save()

                messages.success(request, "Firm saved successfully!")

                # check return target
                next_page = request.GET.get('next')
                if next_page == 'sale':
                    return redirect(reverse('sale_form_new'))
                elif next_page == 'purchase':
                    return redirect(reverse('purchase_form_new'))
                elif next_page == 'daily':
                    return redirect(reverse('daily_page'))

                # default: stay on firm list but provide created_id/name for JS
                created_name = getattr(obj, 'firmname', getattr(obj, 'name', ''))
                return redirect(f"{reverse('firm')}?created_id={obj.pk}&created_name={quote(created_name)}")

        elif request.POST.get('action') == 'delete':
            name = request.POST.get('firmname') or request.POST.get('firm')
            if not name:
                messages.error(request, "No firm specified to delete.")
                return redirect('firm')

            try:
                obj = Firm.objects.get(firmname=name)
                obj.delete()
                messages.success(request, "Firm deleted successfully!")
                return redirect('firm')
            except Firm.DoesNotExist:
                messages.error(request, "Firm not found!")

    return render(request, 'brokerapp/firm.html', {
        'form': form,
        'firms': firms
    })



@require_POST
def firm_delete(request, pk):
    firm = get_object_or_404(Firm, pk=pk)
    firm.delete()
    messages.success(request, f"✅ Firm '{pk}' deleted.")
    return redirect(reverse('firm_list'))

@login_required
def dashboard(request):
    return render(request, 'brokerapp/dashboard.html')


def item_view(request, pk=None):
    # Only fetch inside current org (same as party_view / broker_view)
    assert getattr(request, "current_org", None) is not None, "current_org missing"
    instance = None
    if pk:
        instance = get_object_or_404(HeadItem, pk=pk, org=request.current_org)

    if request.method == 'POST':
        action = request.POST.get('action')

        # ---- DELETE action ----
        if action == 'delete':
            # item_pk may be item_name (string PK) or provided via URL pk
            item_pk = request.POST.get('item_pk') or pk
            if not item_pk:
                messages.error(request, '❌ No item selected to delete.')
                return redirect(reverse('item'))

            try:
                obj_to_delete = get_object_or_404(HeadItem, pk=item_pk, org=request.current_org)
                obj_to_delete.delete()
                messages.success(request, '✅ Item deleted successfully!')
            except Exception as e:
                # catch FK/constraint errors or unexpected issues
                messages.error(request, f'❌ Could not delete item: {e}')
            return redirect(reverse('item'))

        # ---- SAVE / CREATE / UPDATE ----
        form = ItemForm(request.POST, instance=instance, current_org=request.current_org)

        if form.is_valid():
            obj = form.save(commit=False)
            # Ensure item always belongs to selected org
            obj.org = request.current_org
            obj.save()

            if pk:
                messages.success(request, '✅ Item updated successfully!')
            else:
                messages.success(request, '✅ Item added successfully!')

            # Redirect back to Sale if requested
            next_page = request.GET.get('next')
            if next_page == 'sale':
                return redirect(reverse('sale_form_new'))
            elif next_page == 'purchase':
                return redirect(reverse('purchase_form_new'))
            elif next_page == 'daily':
                return redirect(reverse('daily_page'))

            # Default: stay on Item list and pass created id/name for JS
            return redirect(f"{reverse('item')}?created_id={obj.pk}&created_name={quote(obj.item_name)}")
        else:
            messages.error(request, f"❌ Could not save item:\n{form.errors.as_text()}")
    else:
        # pass current_org into form on GET as well
        form = ItemForm(instance=instance, current_org=request.current_org)

    # Show only current org items
    items = HeadItem.objects.filter(org=request.current_org)

    return render(request, 'brokerapp/item.html', {
        'form': form,
        'items': items,
        'editing': pk is not None,
        'editing_id': pk
    })




@require_GET

def daily_page_view(request):
    """
    Show daily page for selected date (via ?date=YYYY-MM-DD) scoped to current org.
    If no date provided, default to today.
    """
    date_str = request.GET.get('date', '').strip()
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()

    # Only current org choices
    parties = HeadParty.objects.filter(org=request.current_org).order_by('partyname')
    brokers = Broker.objects.filter(org=request.current_org).order_by('brokername')

    # Firms for dropdown (scope to org if Firm has org FK, otherwise use all)
    try:
        firm_field_names = [f.name for f in Firm._meta.get_fields()]
        if 'org' in firm_field_names:
            firms = Firm.objects.filter(org=request.current_org).order_by('firmname')
        else:
            firms = Firm.objects.all().order_by('firmname')
    except Exception:
        firms = Firm.objects.all().order_by('firmname')

    # DailyPage for current org + selected date
    daily_page = DailyPage.objects.filter(
        org=request.current_org,
        date=selected_date
    ).first()

    def serialize_entry(e):
        # party_name (works if e.party is FK or string)
        party_name = ''
        if getattr(e, 'party', None):
            party_name = getattr(getattr(e, 'party', None), 'partyname', None) or getattr(e, 'party', None) or ''
        else:
            party_name = getattr(e, 'party_name', '') or ''

        # broker_name (works if e.broker is FK or string)
        broker_name = ''
        if getattr(e, 'broker', None):
            broker_name = getattr(getattr(e, 'broker', None), 'brokername', None) or getattr(e, 'broker', None) or ''
        else:
            broker_name = getattr(e, 'broker_name', '') or ''

        # firm_name (works if e.firm is FK or there is a firm_name string field)
        firm_name = ''
        if getattr(e, 'firm', None):
            firm_name = getattr(getattr(e, 'firm', None), 'firmname', None) or getattr(e, 'firm', None) or ''
        else:
            firm_name = getattr(e, 'firm_name', '') or ''

        return {
            'entry_no': e.entry_no,
            'party_name': party_name,
            'broker_name': broker_name,
            'firm_name': firm_name,
            'amount': e.amount,
            'remark': e.remark,
        }

    jama_entries = []
    naame_entries = []

    if daily_page:
        # Build safe select_related list based on actual fields on the entry model
        try:
            jama_model = daily_page.jama_entries.model
            naame_model = daily_page.naame_entries.model

            def safe_select_qs(qs, model):
                field_names = [f.name for f in model._meta.get_fields()]
                rels = [name for name in ('party', 'broker', 'firm') if name in field_names]
                if rels:
                    return qs.select_related(*rels).all()
                return qs.all()

            jama_qs = safe_select_qs(daily_page.jama_entries, jama_model)
            naame_qs = safe_select_qs(daily_page.naame_entries, naame_model)
        except Exception:
            # Fallback: no select_related
            jama_qs = daily_page.jama_entries.all()
            naame_qs = daily_page.naame_entries.all()

        jama_entries = [serialize_entry(j) for j in jama_qs]
        naame_entries = [serialize_entry(n) for n in naame_qs]

    no_entries = not (jama_entries or naame_entries)

    context = {
        'selected_date': selected_date,
        'parties': parties,
        'brokers': brokers,
        'firms': firms,
        'jama_entries': jama_entries,
        'naame_entries': naame_entries,
        'no_entries': no_entries,
    }
    return render(request, 'brokerapp/daily_page.html', context)

@require_GET

def daily_page_show(request):
    """
    JSON endpoint: ?date=YYYY-MM-DD (optional; if missing -> today)
    Response: { "date": "...", "jama": [...], "naame": [...] }
    """
    d = request.GET.get('date', '').strip()

    # default to today if missing or invalid
    if not d:
        date_obj = timezone.localdate()
    else:
        try:
            date_obj = datetime.strptime(d, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'invalid date format, expected YYYY-MM-DD'}, status=400)

    # scope to current org
    daily_page = DailyPage.objects.filter(org=request.current_org, date=date_obj).first()

    def serialize_entry(entry):
        # Party
        if getattr(entry, 'party', None):
            party_name = getattr(getattr(entry, 'party', None), 'partyname', None) or str(entry.party)
        else:
            party_name = getattr(entry, 'party_name', '') or ''

        # Broker
        if getattr(entry, 'broker', None):
            broker_name = getattr(getattr(entry, 'broker', None), 'brokername', None) or str(entry.broker)
        else:
            broker_name = getattr(entry, 'broker_name', '') or ''

        # Firm (works for FK `firm` or string `firm_name`)
        if getattr(entry, 'firm', None):
            firm_name = getattr(getattr(entry, 'firm', None), 'firmname', None) or str(entry.firm)
        else:
            firm_name = getattr(entry, 'firm_name', '') or ''

        return {
            'entry_no': entry.entry_no,
            'party_name': party_name,
            'broker_name': broker_name,
            'firm_name': firm_name,
            'amount': float(entry.amount or 0),
            'remark': entry.remark or '',
            'created_at': entry.created_at.isoformat() if getattr(entry, 'created_at', None) else None,
        }

    jama = []
    naame = []

    if daily_page:
        # Build safe select_related based on the actual fields on the entry models
        try:
            jama_model = daily_page.jama_entries.model
            naame_model = daily_page.naame_entries.model

            def safe_qs(qs, model):
                field_names = [f.name for f in model._meta.get_fields()]
                rels = [name for name in ('party', 'broker', 'firm') if name in field_names]
                if rels:
                    return qs.select_related(*rels).all()
                return qs.all()

            jama_qs = safe_qs(daily_page.jama_entries, jama_model)
            naame_qs = safe_qs(daily_page.naame_entries, naame_model)
        except Exception:
            jama_qs = daily_page.jama_entries.all()
            naame_qs = daily_page.naame_entries.all()

        jama = [serialize_entry(j) for j in jama_qs]
        naame = [serialize_entry(n) for n in naame_qs]

    if not jama and not naame:
        return JsonResponse({
            'date': date_obj.strftime('%Y-%m-%d'),
            'message': 'No entry on that day',
            'jama': [],
            'naame': [],
        })

    return JsonResponse({
        'date': date_obj.strftime('%Y-%m-%d'),
        'jama': jama,
        'naame': naame,
    })

    
@require_POST

def daily_page_jama_add(request):
    # expects: date, party (pk), broker (pk), firm (pk, optional), amount, remark (optional)
    date_str = request.POST.get('date')
    party_id = request.POST.get('party')
    broker_id = request.POST.get('broker')
    firm_id = request.POST.get('firm')   # NEW: optional
    amount = request.POST.get('amount')
    remark = (request.POST.get('remark') or '').strip()

    if not (date_str and party_id and amount):
        return JsonResponse({'error': 'Missing required fields (date/party/amount)'}, status=400)

    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        amt = float(amount)
    except Exception:
        return JsonResponse({'error': 'Invalid input'}, status=400)

    # Party/Broker must belong to current org (broker optional if your UI allows blank)
    party = get_object_or_404(HeadParty, pk=party_id, org=request.current_org)
    broker = None
    if broker_id:
        broker = get_object_or_404(Broker, pk=broker_id, org=request.current_org)

    # Resolve firm (optional). Be defensive: Firm model may or may not have 'org' field.
    firm_obj = None
    firm_name_text = ''
    if firm_id:
        try:
            firm_field_names = [f.name for f in Firm._meta.get_fields()]
            if 'org' in firm_field_names:
                firm_obj = get_object_or_404(Firm, pk=firm_id, org=request.current_org)
            else:
                firm_obj = get_object_or_404(Firm, pk=firm_id)
            firm_name_text = getattr(firm_obj, 'firmname', str(firm_obj))
        except Exception:
            firm_obj = None
            firm_name_text = firm_id  # fallback to posted value

    with transaction.atomic():
        # DailyPage is per (org, date)
        daily_page, _ = DailyPage.objects.get_or_create(org=request.current_org, date=date_obj)

        # Prepare entry kwargs dynamically (support both FK 'firm' or string 'firm_name')
        entry_kwargs = {
            'daily_page': daily_page,
            'party': party,
            'amount': amt,
            'remark': remark,
        }
        if broker is not None:
            entry_kwargs['broker'] = broker

        # Detect JamaEntry fields
        try:
            entry_field_names = [f.name for f in JamaEntry._meta.get_fields()]
        except Exception:
            entry_field_names = []

        if firm_id:
            if 'firm' in entry_field_names:
                entry_kwargs['firm'] = firm_obj  # may be None if resolution failed
            elif 'firm_name' in entry_field_names:
                entry_kwargs['firm_name'] = firm_name_text
            # otherwise ignore if neither field present

        entry = JamaEntry.objects.create(**entry_kwargs)

    data = {
        'entry_no': entry.entry_no,
        'party_name': party.partyname,
        'broker_name': broker.brokername if broker else '',
        'firm_name': firm_name_text,
        'amount': f"{entry.amount:.2f}",
        'remark': entry.remark,
    }
    return JsonResponse({'success': True, 'entry': data})


@require_POST

def daily_page_naame_add(request):
    # expects: date, party (pk), broker (pk), firm (pk, optional), amount, remark (optional)
    date_str = request.POST.get('date')
    party_id = request.POST.get('party')
    broker_id = request.POST.get('broker')
    firm_id = request.POST.get('firm')   # NEW: optional
    amount = request.POST.get('amount')
    remark = (request.POST.get('remark') or '').strip()

    if not (date_str and party_id and amount):
        return JsonResponse({'error': 'Missing required fields (date/party/amount)'}, status=400)

    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        amt = float(amount)
    except Exception:
        return JsonResponse({'error': 'Invalid input'}, status=400)

    # Scoped lookups
    party = get_object_or_404(HeadParty, pk=party_id, org=request.current_org)
    broker = None
    if broker_id:
        broker = get_object_or_404(Broker, pk=broker_id, org=request.current_org)

    # Resolve firm (optional) - defensive about Firm model shape
    firm_obj = None
    firm_name_text = ''
    if firm_id:
        try:
            firm_field_names = [f.name for f in Firm._meta.get_fields()]
            if 'org' in firm_field_names:
                firm_obj = get_object_or_404(Firm, pk=firm_id, org=request.current_org)
            else:
                firm_obj = get_object_or_404(Firm, pk=firm_id)
            firm_name_text = getattr(firm_obj, 'firmname', str(firm_obj))
        except Exception:
            firm_obj = None
            firm_name_text = firm_id

    with transaction.atomic():
        daily_page, _ = DailyPage.objects.get_or_create(org=request.current_org, date=date_obj)

        entry_kwargs = {
            'daily_page': daily_page,
            'party': party,
            'amount': amt,
            'remark': remark,
        }
        if broker is not None:
            entry_kwargs['broker'] = broker

        # Detect NaameEntry fields (support FK 'firm' or string 'firm_name')
        try:
            entry_field_names = [f.name for f in NaameEntry._meta.get_fields()]
        except Exception:
            entry_field_names = []

        if firm_id:
            if 'firm' in entry_field_names:
                entry_kwargs['firm'] = firm_obj
            elif 'firm_name' in entry_field_names:
                entry_kwargs['firm_name'] = firm_name_text

        entry = NaameEntry.objects.create(**entry_kwargs)

    data = {
        'entry_no': entry.entry_no,
        'party_name': party.partyname,
        'broker_name': broker.brokername if broker else '',
        'firm_name': firm_name_text,
        'amount': f"{entry.amount:.2f}",
        'remark': entry.remark,
    }
    return JsonResponse({'success': True, 'entry': data})

@require_POST
def daily_page_jama_delete(request, entry_no):
    entry = get_object_or_404(JamaEntry, entry_no=entry_no, daily_page__org=request.current_org)
    entry.delete()
    return JsonResponse({'success': True, 'entry_no': entry_no})

@require_POST
def daily_page_jama_update(request):
    entry_no = request.POST.get("entry_no")
    if not entry_no:
        return JsonResponse({"error": "entry_no required"}, status=400)

    entry = get_object_or_404(JamaEntry, entry_no=entry_no, daily_page__org=request.current_org)

    try:
        with transaction.atomic():
            # basic fields
            if request.POST.get("amount") not in (None, ''):
                entry.amount = request.POST.get("amount")
            entry.remark = request.POST.get("remark", "") or ""

            # party: if user supplied text, store into party_name and clear FK; if they selected FK (id), try to set FK
            party_val = request.POST.get("party")
            # if the select in your template sends a partyname string (not id), we treat it as text
            # if you later change frontend to send party_id, adapt here.
            if party_val:
                # if front-end sends the partyname string (as earlier), store it in party_name and clear FK:
                entry.party_name = party_val
                entry.party = None
            # broker: treat as name string (same approach)
            broker_val = request.POST.get("broker")
            if broker_val:
                entry.broker = None
                # ensure text fallback (broker name field may be named differently in models — using broker_name if present)
                if hasattr(entry, 'broker_name'):
                    entry.broker_name = broker_val
                elif hasattr(entry, 'brokername'):
                    entry.brokername = broker_val

            # firm: accept id or textual; try to treat as id first
            firm_val = request.POST.get("firm")
            if firm_val:
                try:
                    entry.firm_id = int(firm_val)
                    # refresh firm_name from FK if available
                    entry.firm_name = getattr(entry.firm, 'firmname', '') or entry.firm_name
                except Exception:
                    entry.firm = None
                    entry.firm_id = None
                    entry.firm_name = firm_val or entry.firm_name

            entry.save()
    except Exception as exc:
        return JsonResponse({"error": "exception saving entry: " + str(exc)}, status=500)

    return JsonResponse({
        "success": True,
        "entry": {
            "entry_no": entry.entry_no,
            "amount": str(entry.amount),
            "party_name": entry.party_name or "",
            "broker_name": getattr(entry, 'broker_name', '') or "",
            "firm_name": entry.firm_name or "",
            "remark": entry.remark or "",
        }
    })

    
@require_POST
def daily_page_naame_delete(request, entry_no):
    entry = get_object_or_404(NaameEntry, entry_no=entry_no, daily_page__org=request.current_org)
    entry.delete()
    return JsonResponse({'success': True, 'entry_no': entry_no})


@require_POST
def daily_page_naame_update(request):
    entry_no = request.POST.get("entry_no")
    if not entry_no:
        return JsonResponse({"error": "entry_no required"}, status=400)

    entry = get_object_or_404(NaameEntry, entry_no=entry_no, daily_page__org=request.current_org)

    try:
        with transaction.atomic():
            if request.POST.get("amount") not in (None, ''):
                entry.amount = request.POST.get("amount")
            entry.remark = request.POST.get("remark", "") or ""

            party_val = request.POST.get("party")
            if party_val:
                entry.party_name = party_val
                entry.party = None

            broker_val = request.POST.get("broker")
            if broker_val:
                entry.broker = None
                if hasattr(entry, 'broker_name'):
                    entry.broker_name = broker_val
                elif hasattr(entry, 'brokername'):
                    entry.brokername = broker_val

            firm_val = request.POST.get("firm")
            if firm_val:
                try:
                    entry.firm_id = int(firm_val)
                    entry.firm_name = getattr(entry.firm, 'firmname', '') or entry.firm_name
                except Exception:
                    entry.firm = None
                    entry.firm_id = None
                    entry.firm_name = firm_val or entry.firm_name

            entry.save()
    except Exception as exc:
        return JsonResponse({"error": "exception saving entry: " + str(exc)}, status=500)

    return JsonResponse({
        "success": True,
        "entry": {
            "entry_no": entry.entry_no,
            "amount": str(entry.amount),
            "party_name": entry.party_name or "",
            "broker_name": getattr(entry, 'broker_name', '') or "",
            "firm_name": entry.firm_name or "",
            "remark": entry.remark or "",
        }
    })
    
def daily_page_pdf(request):
    date = request.GET.get('date')
    if not date:
        return HttpResponse("Date not provided", status=400)

    # Query entries (keeps original filtering by date)
    jama_entries = list(JamaEntry.objects.filter(daily_page__date=date).order_by('entry_no'))
    naame_entries = list(NaameEntry.objects.filter(daily_page__date=date).order_by('entry_no'))

    total_jama = sum(float(j.amount or 0) for j in jama_entries)
    total_naame = sum(float(n.amount or 0) for n in naame_entries)
    diff = total_jama - total_naame

    # --- PDF setup (landscape A4) ---
    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, f"Daily Report - {date}", ln=True, align="C")
    pdf.ln(6)

    # Column layout (left and right table)
    left_x = 12
    right_x = 12 + 137 + 8  # left panel width + small gap
    panel_width = 137  # width for each table panel

    # Column widths inside a panel (sum <= panel_width)
    # no, party, broker, firm, amount, remark
    w_no = 10
    w_party = 40
    w_broker = 30
    w_firm = 25
    w_amount = 20
    w_remark = panel_width - (w_no + w_party + w_broker + w_firm + w_amount)

    # Header row height
    hdr_h = 8
    row_h = 8

    # Helper: safe text for party/broker/firm (handles FK or stored-name)
    def txt_party(e):
        return getattr(getattr(e, 'party', None), 'partyname', None) or getattr(e, 'party_name', None) or (str(e.party) if getattr(e, 'party', None) else '')

    def txt_broker(e):
        return getattr(getattr(e, 'broker', None), 'brokername', None) or getattr(e, 'broker_name', None) or (str(e.broker) if getattr(e, 'broker', None) else '')

    def txt_firm(e):
        return getattr(getattr(e, 'firm', None), 'firmname', None) or getattr(e, 'firm_name', None) or (str(e.firm) if getattr(e, 'firm', None) else '')

    # Function to draw a single panel (list of entries)
    def draw_panel(x, y, title, entries):
        pdf.set_xy(x, y)
        pdf.set_font("Arial", "B", 12)
        pdf.cell(panel_width, 6, title, ln=1)

        # table header
        pdf.set_font("Arial", "B", 10)
        pdf.set_xy(x, pdf.get_y())
        pdf.cell(w_no, hdr_h, "No", border=1, align="C")
        pdf.cell(w_party, hdr_h, "Party", border=1, align="L")
        pdf.cell(w_broker, hdr_h, "Broker", border=1, align="L")
        pdf.cell(w_firm, hdr_h, "Firm", border=1, align="L")
        pdf.cell(w_amount, hdr_h, "Amount", border=1, align="R")
        pdf.cell(w_remark, hdr_h, "Remark", border=1, align="L")
        pdf.ln(hdr_h)

        # rows
        pdf.set_font("Arial", "", 10)
        for e in entries:
            pdf.set_x(x)
            pdf.cell(w_no, row_h, str(e.entry_no), border=1)
            # party (truncate if too long)
            party_text = txt_party(e) or ""
            if len(party_text) > 30:
                party_text = party_text[:27] + "..."
            pdf.cell(w_party, row_h, party_text, border=1)
            # broker
            broker_text = txt_broker(e) or ""
            if len(broker_text) > 25:
                broker_text = broker_text[:22] + "..."
            pdf.cell(w_broker, row_h, broker_text, border=1)
            # firm
            firm_text = txt_firm(e) or ""
            if len(firm_text) > 25:
                firm_text = firm_text[:22] + "..."
            pdf.cell(w_firm, row_h, firm_text, border=1)
            # amount (right aligned)
            pdf.cell(w_amount, row_h, f"{float(e.amount or 0):.2f}", border=1, align="R")
            # remark - truncate
            remark_text = (e.remark or "")
            if len(remark_text) > 30:
                remark_text = remark_text[:27] + "..."
            pdf.cell(w_remark, row_h, remark_text, border=1)
            pdf.ln(row_h)

        # Totals row
        pdf.set_x(x)
        span_width = w_no + w_party + w_broker + w_firm
        pdf.set_font("Arial", "B", 10)
        pdf.cell(span_width, hdr_h, "Total", border='T')
        panel_total = sum(float(ent.amount or 0) for ent in entries)
        pdf.cell(w_amount, hdr_h, f"{panel_total:.2f}", border='T', align="R")
        pdf.cell(w_remark, hdr_h, "", border='T')
        pdf.ln(hdr_h + 4)

    # Draw both panels side-by-side starting from current y
    start_y = pdf.get_y()
    draw_panel(left_x, start_y, "Jama", jama_entries)
    draw_panel(right_x, start_y, "Naame", naame_entries)

    # Summary line (below panels)
    pdf.set_font("Arial", "B", 12)
    pdf.set_xy(left_x, pdf.get_y())
    pdf.cell(120, 8, f"Jama Total: {total_jama:.2f}", ln=0)
    pdf.set_xy(right_x, pdf.get_y())
    pdf.cell(120, 8, f"Naame Total: {total_naame:.2f}", ln=0)
    pdf.ln(10)

    # Difference on the right aligned
    pdf.set_font("Arial", "B", 12)
    page_right = pdf.w - 12
    diff_text = f"Difference (Jama - Naame): {diff:.2f}"
    text_width = pdf.get_string_width(diff_text) + 2
    pdf.set_xy(page_right - text_width, pdf.get_y())
    pdf.cell(text_width, 8, diff_text, ln=1, align='R')

    # Output and return
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="DailyReport_{date}.pdf"'
    return response





class AllPartyBalanceView(TemplateView):
    """
    Simplified AllPartyBalanceView matching compact template (no filters).
    Supports POST actions via buttons with name="action":
      - balance       : show table in page
      - print         : render printable HTML (user can browser-print)
      - export_excel  : return .xlsx (requires openpyxl)
      - pdf           : return PDF generated with fpdf2
    """
    template_name = "brokerapp/account/all_party_balance.html"
    printable_template = "brokerapp/account/all_party_balance_printable.html"

    # ---------- helpers ----------
    def _org_filter(self, qs):
        org_id = self.request.session.get("org_id")
        if not org_id:
            return qs
        field_names = [f.attname for f in qs.model._meta.fields]
        if "org_id" in field_names:
            return qs.filter(org_id=org_id)
        return qs

    def _sum(self, qs, field):
        """Safe sum returning Decimal(0) when None."""
        return qs.aggregate(t=Sum(field))["t"] or Decimal("0")

    # ---------- GET ----------
    def get(self, request, *args, **kwargs):
        today = date.today()
        ctx = self._build_context(start=today, end=today, party=None)
        ctx["show_table"] = False
        return self.render_to_response(ctx)

    # ---------- POST ----------
    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        today = date.today()

        # Build rows/totals (same data used by all actions)
        ctx = self._build_context(start=today, end=today, party=None)

        # Balance -> show table in same template
        if action == "balance" or not action:
            ctx["show_table"] = True
            return self.render_to_response(ctx)

        # Print -> render printable HTML (no buttons)
        if action == "print":
            ctx["show_table"] = True
            return render(request, self.printable_template, ctx)

        # Export Excel -> create .xlsx (requires openpyxl)
        if action == "export_excel":
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except Exception:
                return HttpResponse(
                    "Required package 'openpyxl' not installed. Install with: pip install openpyxl",
                    content_type="text/plain",
                    status=500
                )

            wb = Workbook()
            ws = wb.active
            ws.title = "All Party Balance"

            headers = ["Party", "Op Dr", "Op Cr", "Opening", "Sale", "Purchase", "Naame", "Jama", "Balance"]
            ws.append(headers)

            for r in ctx["rows"]:
                pname = getattr(r["party"], "partyname", str(r["party"]))
                ws.append([
                    pname,
                    float(r["op_dr"]), float(r["op_cr"]),
                    float(r["opening"]), float(r["sale"]),
                    float(r["purchase"]), float(r["naame"]),
                    float(r["jama"]), float(r["balance"])
                ])

            # auto column width (simple)
            for i, col in enumerate(ws.columns, start=1):
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            resp = HttpResponse(
                out.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = f'attachment; filename="all_party_balance_{today}.xlsx"'
            return resp

        # PDF -> generate using fpdf2
        if action == "pdf":
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=10)

            # Header
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, "All Party Balance", ln=True, align="C")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 6, f"Generated on: {today.strftime('%d-%m-%Y')}", ln=True, align="C")
            pdf.ln(4)

            # Table headers
            headers = ["Party", "Op Dr", "Op Cr", "Opening", "Sale", "Purchase", "Naame", "Jama", "Balance"]
            col_widths = [50, 18, 18, 24, 18, 22, 18, 18, 22]  # total should fit A4 width with margins

            pdf.set_font("Helvetica", "B", 9)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align="C")
            pdf.ln(8)

            # Rows
            pdf.set_font("Helvetica", "", 9)
            for r in ctx["rows"]:
                vals = [
                    getattr(r["party"], "partyname", str(r["party"])),
                    f"{r['op_dr']:.2f}", f"{r['op_cr']:.2f}",
                    f"{r['opening']:.2f}", f"{r['sale']:.2f}",
                    f"{r['purchase']:.2f}", f"{r['naame']:.2f}",
                    f"{r['jama']:.2f}", f"{r['balance']:.2f}"
                ]
                for i, v in enumerate(vals):
                    align = "L" if i == 0 else "R"
                    pdf.cell(col_widths[i], 7, v, border=1, align=align)
                pdf.ln(7)

            # Totals row
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(col_widths[0], 8, "TOTAL", border=1, align="L")
            totals = ctx["totals"]
            total_vals = [
                totals["opdr"], totals["opcr"], totals["sale"], totals["purchase"],
                totals["naame"], totals["jama"], totals["balance"]
            ]
            # place totals aligned under numeric columns (skip party col)
            # mapping to header indices: 1(OpDr),2(OpCr),3(Opening) etc. We'll print totals aligned with numeric columns.
            # For simplicity, print totals under Op Dr onward; keep Opening blank since it's derived per-party.
            pdf.cell(col_widths[1], 8, f"{totals['opdr']:.2f}", border=1, align="R")
            pdf.cell(col_widths[2], 8, f"{totals['opcr']:.2f}", border=1, align="R")
            pdf.cell(col_widths[3], 8, "", border=1, align="R")  # Opening total left blank
            pdf.cell(col_widths[4], 8, f"{totals['sale']:.2f}", border=1, align="R")
            pdf.cell(col_widths[5], 8, f"{totals['purchase']:.2f}", border=1, align="R")
            pdf.cell(col_widths[6], 8, f"{totals['naame']:.2f}", border=1, align="R")
            pdf.cell(col_widths[7], 8, f"{totals['jama']:.2f}", border=1, align="R")
            pdf.cell(col_widths[8], 8, f"{totals['balance']:.2f}", border=1, align="R")
            pdf.ln(10)

            buf = io.BytesIO()
            pdf.output(buf)
            buf.seek(0)
            resp = HttpResponse(buf.read(), content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="all_party_balance_{today}.pdf"'
            return resp

        # Unknown action -> render without table
        ctx["show_table"] = False
        return self.render_to_response(ctx)

    # ---------- core calculation ----------
    def _build_context(self, start, end, party):
        # parties
        parties = HeadParty.objects.all().order_by("partyname")
        if party:
            parties = parties.filter(partyname=party.partyname)

        rows = []
        totals = {
            "opdr": Decimal("0"), "opcr": Decimal("0"),
            "sale": Decimal("0"), "purchase": Decimal("0"),
            "naame": Decimal("0"), "jama": Decimal("0"),
            "balance": Decimal("0")
        }

        dp_before = DailyPage.objects.filter(date__lt=start)
        dp_range = DailyPage.objects.filter(date__range=(start, end))

        org_id = self.request.session.get("org_id")
        if org_id:
            dp_before = dp_before.filter(org_id=org_id)
            dp_range = dp_range.filter(org_id=org_id)
            parties = parties.filter(org_id=org_id)

        for p in parties:
            op_dr = Decimal(getattr(p, "openingdebit", 0) or 0)
            op_cr = Decimal(getattr(p, "openingcredit", 0) or 0)

            sale_before = self._sum(self._org_filter(SaleMaster.objects.filter(party=p, invdate__lt=start)), "netamt")
            purch_before = self._sum(self._org_filter(PurchaseMaster.objects.filter(party=p, invdate__lt=start)), "netamt")
            naame_before = self._sum(NaameEntry.objects.filter(daily_page__in=dp_before, party=p), "amount")
            jama_before = self._sum(JamaEntry.objects.filter(daily_page__in=dp_before, party=p), "amount")

            opening = (op_dr - op_cr) + (sale_before - purch_before + naame_before - jama_before)

            sale = self._sum(self._org_filter(SaleMaster.objects.filter(party=p, invdate__range=(start, end))), "netamt")
            purchase = self._sum(self._org_filter(PurchaseMaster.objects.filter(party=p, invdate__range=(start, end))), "netamt")
            naame = self._sum(NaameEntry.objects.filter(daily_page__in=dp_range, party=p), "amount")
            jama = self._sum(JamaEntry.objects.filter(daily_page__in=dp_range, party=p), "amount")

            balance = opening + sale - purchase + naame - jama

            rows.append({
                "party": p, "op_dr": op_dr, "op_cr": op_cr, "opening": opening,
                "sale": sale, "purchase": purchase, "naame": naame,
                "jama": jama, "balance": balance
            })

            totals["opdr"] += op_dr
            totals["opcr"] += op_cr
            totals["sale"] += sale
            totals["purchase"] += purchase
            totals["naame"] += naame
            totals["jama"] += jama
            totals["balance"] += balance

        return {"rows": rows, "totals": totals, "start": start, "end": end}

# ---------- original party_statement (uses helper) ----------
class PartyStatementView(TemplateView):
    """
    Single URL Party Statement view. Buttons POST with name="action":
     - statement      : show table
     - print          : render printable HTML
     - export_excel   : return .xlsx
     - pdf            : return PDF (fpdf)
    """
    template_name = "brokerapp/account/party_statement.html"
    printable_template = "brokerapp/account/party_statement_printable.html"

    def get(self, request, *args, **kwargs):
        parties = HeadParty.objects.order_by("partyname")
        ctx = {
            "parties": parties,
            "selected": None,
            "entries": [],
            "total_debit": Decimal("0"),
            "total_credit": Decimal("0"),
            "balance": Decimal("0"),
        }
        return self.render_to_response(ctx)

    def post(self, request, *args, **kwargs):
        """
        Process POST actions. Always returns an HttpResponse.
        """
        action = request.POST.get("action")
        # handle both header POST (party in POST) or fallback to GET param
        party_id = request.POST.get("party") or request.GET.get("party")
        parties = HeadParty.objects.order_by("partyname")

        # if no party selected and action requires party -> show page with message
        if not party_id:
            # For actions that do not require a party, still return page (here all require a party)
            ctx = {
                "parties": parties,
                "selected": None,
                "entries": [],
                "total_debit": Decimal("0"),
                "total_credit": Decimal("0"),
                "balance": Decimal("0"),
            }
            return self.render_to_response(ctx)

        # load party and compute entries
        head = get_object_or_404(HeadParty, pk=party_id)
        entries, total_debit, total_credit, balance = self._build_entries(head)

        ctx = {
            "parties": parties,
            "selected": head,
            "entries": entries,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balance": balance,
            "today": date.today(),
        }

        # ---------- show statement ----------
        if action in (None, "statement"):
            return self.render_to_response(ctx)

        
        # ---------- email (PDF attachment via FPDF) ----------
        if action == "email":
            # party ke paas email na ho to warning
            if not getattr(head, "email", None):
                messages.warning(request, "Selected party has no email address.")
                return self.render_to_response(ctx)

            if FPDF is None:
                messages.warning(
                    request,
                    "PDF package 'fpdf' not installed. Install with: pip install fpdf",
                )
                return self.render_to_response(ctx)

            # helper to avoid FPDF unicode errors (keeps ascii only)
            def safe_text(val, maxlen=None):
                s = "" if val is None else str(val)
                s = s.replace("—", "-").replace("–", "-")
                s = "".join(ch if ord(ch) < 128 else "?" for ch in s)
                return s[:maxlen] if maxlen else s

            try:
                # --- 1) FPDF PDF generate (same style as pdf action) ---
                pdf = FPDF()
                pdf.add_page()
                pdf.set_auto_page_break(auto=True, margin=10)

                # Header
                pdf.set_font("Helvetica", "B", 14)
                pdf.cell(0, 10, safe_text(f"Party Statement - {head.partyname}", 140), ln=True, align="C")
                pdf.set_font("Helvetica", "", 10)
                pdf.cell(0, 6, safe_text(f"Generated on: {date.today().strftime('%d-%m-%Y')}", 80), ln=True, align="C")
                pdf.ln(4)

                headers = ["Entry No", "Date", "Debit", "Credit", "Firm", "Remark", "Balance"]
                widths = widths = [18, 20, 22, 22, 24, 60, 24]


                pdf.set_font("Helvetica", "B", 9)
                for i, h in enumerate(headers):
                    pdf.cell(widths[i], 8, safe_text(h, 40), border=1, align="C")
                pdf.ln(8)

                pdf.set_font("Helvetica", "", 9)
                for e in entries:
                    vals = [
                        safe_text(e.get("entry_no", ""), 20),
                        safe_text(e["date"].strftime("%Y-%m-%d") if e["date"] else "", 20),
                        safe_text(f"{(e.get('debit') or Decimal('0')):.2f}", 20),
                        safe_text(f"{(e.get('credit') or Decimal('0')):.2f}", 20),
                        safe_text(e.get("firm_name", ""), 28),
                        safe_text(e.get("remark", ""), 120),
                        safe_text(f"{(e.get('balance') or Decimal('0')):.2f}", 20),
                    ]
                    for i, v in enumerate(vals):
                        pdf.cell(widths[i], 7, v, border=1, align="L" if i in (0, 1, 4) else "R")
                    pdf.ln(7)

                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(widths[0] + widths[1], 8, safe_text("TOTAL", 40), border=1, align="L")
                pdf.cell(widths[2], 8, safe_text(f"{total_debit:.2f}", 20), border=1, align="R")
                pdf.cell(widths[3], 8, safe_text(f"{total_credit:.2f}", 20), border=1, align="R")
                pdf.cell(widths[4], 8, "", border=1, align="R")   # Firm blank
                pdf.cell(widths[5], 8, "", border=1, align="R")   # Remark blank
                pdf.cell(widths[6], 8, safe_text(f"{balance:.2f}", 20), border=1, align="R")  # Balance total


                buf = io.BytesIO()
                pdf.output(buf)
                buf.seek(0)
                pdf_bytes = buf.read()

                safe_name = "".join(ch if ord(ch) < 128 else "?" for ch in head.partyname)[:40]

                # --- 2) Email with PDF attachment ---
                subject = f"Party Statement - {head.partyname}"
                body = (
                    f"Dear {head.partyname},\n\n"
                    "Please find your detailed party statement attached as PDF.\n\n"
                    f"Total Debit : {total_debit:.2f}\n"
                    f"Total Credit: {total_credit:.2f}\n"
                    f"Balance     : {balance:.2f}\n\n"
                    "Thank you."
                )

                email = EmailMessage(
                    subject,
                    body,
                    settings.DEFAULT_FROM_EMAIL,
                    [head.email],
                )
                email.attach(
                    f"party_statement_{safe_name}.pdf",
                    pdf_bytes,
                    "application/pdf",
                )
                email.send(fail_silently=False)

                messages.success(request, f"Party statement PDF emailed to {head.email}.")

            except Exception as e:
                messages.warning(request, f"Statement PDF ready but email not sent: {e}")

            return self.render_to_response(ctx)

        
        # ---------- printable ----------
        if action == "print":
            return render(request, self.printable_template, ctx)

        # ---------- excel ----------
        if action == "export_excel":
            if Workbook is None:
                return HttpResponse(
                    "Required package 'openpyxl' not installed. Install with: pip install openpyxl",
                    content_type="text/plain",
                    status=500
                )
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Party Statement"
                headers = ["Entry No", "Date", "Debit", "Credit", "Firm", "Remark", "Balance"]
                ws.append(headers)
                for e in entries:
                    ws.append([
                        e.get("entry_no"),
                        e["date"].strftime("%Y-%m-%d") if e["date"] else "",
                        float(e.get("debit") or 0),
                        float(e.get("credit") or 0),
                        e.get("firm_name") or "",
                        e.get("remark") or "",
                        float(e.get("balance") or 0),
                    ])
                ws.append([])
                ws.append(["", "Total", float(total_debit), float(total_credit), "", float(balance)])

                if get_column_letter:
                    for i, col in enumerate(ws.columns, start=1):
                        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

                out = io.BytesIO()
                wb.save(out)
                out.seek(0)
                resp = HttpResponse(
                    out.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                safe_name = "".join(ch if ord(ch) < 128 else "?" for ch in head.partyname)[:40]
                resp["Content-Disposition"] = f'attachment; filename="party_statement_{safe_name}.xlsx"'
                return resp
            except Exception as exc:
                return HttpResponse(f"Excel export failed: {exc}", content_type="text/plain", status=500)

        # ---------- pdf ----------
        if action == "pdf":
            if FPDF is None:
                return HttpResponse(
                    "Required package 'fpdf' not installed. Install with: pip install fpdf",
                    content_type="text/plain",
                    status=500
                )

            # helper to avoid FPDF unicode errors (keeps ascii only)
            def safe_text(val, maxlen=None):
                s = "" if val is None else str(val)
                s = s.replace("—", "-").replace("–", "-")
                s = "".join(ch if ord(ch) < 128 else "?" for ch in s)
                return s[:maxlen] if maxlen else s

            try:
                pdf = FPDF()
                pdf.add_page()
                pdf.set_auto_page_break(auto=True, margin=10)

                # Header
                pdf.set_font("Helvetica", "B", 14)
                pdf.cell(0, 10, safe_text(f"Party Statement - {head.partyname}", 140), ln=True, align="C")
                pdf.set_font("Helvetica", "", 10)
                pdf.cell(0, 6, safe_text(f"Generated on: {date.today().strftime('%d-%m-%Y')}", 80), ln=True, align="C")
                pdf.ln(4)

                headers = ["Entry No", "Date", "Debit", "Credit", "Firm", "Remark", "Balance"]
                widths = widths = [18, 20, 22, 22, 24, 60, 24]

                pdf.set_font("Helvetica", "B", 9)
                for i, h in enumerate(headers):
                    pdf.cell(widths[i], 8, safe_text(h, 40), border=1, align="C")
                pdf.ln(8)

                pdf.set_font("Helvetica", "", 9)
                for e in entries:
                    vals = [
                        safe_text(e.get("entry_no", ""), 20),
                        safe_text(e["date"].strftime("%Y-%m-%d") if e["date"] else "", 20),
                        safe_text(f"{(e.get('debit') or Decimal('0')):.2f}", 20),
                        safe_text(f"{(e.get('credit') or Decimal('0')):.2f}", 20),
                        safe_text(e.get("firm_name", ""), 28),
                        safe_text(e.get("remark", ""), 120),
                        safe_text(f"{(e.get('balance') or Decimal('0')):.2f}", 20),
                    ]
                    for i, v in enumerate(vals):
                        pdf.cell(widths[i], 7, v, border=1, align="L" if i in (0, 1, 4) else "R")
                    pdf.ln(7)

                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(widths[0] + widths[1], 8, safe_text("TOTAL", 40), border=1, align="L")
                pdf.cell(widths[2], 8, safe_text(f"{total_debit:.2f}", 20), border=1, align="R")
                pdf.cell(widths[3], 8, safe_text(f"{total_credit:.2f}", 20), border=1, align="R")
                pdf.cell(widths[4], 8, "", border=1, align="R")   # Firm blank
                pdf.cell(widths[5], 8, "", border=1, align="R")   # Remark blank
                pdf.cell(widths[6], 8, safe_text(f"{balance:.2f}", 20), border=1, align="R")  # Balance total


                buf = io.BytesIO()
                pdf.output(buf)
                buf.seek(0)
                resp = HttpResponse(buf.read(), content_type="application/pdf")
                safe_name = "".join(ch if ord(ch) < 128 else "?" for ch in head.partyname)[:40]
                resp["Content-Disposition"] = f'attachment; filename="party_statement_{safe_name}.pdf"'
                return resp
            except Exception as exc:
                return HttpResponse(f"PDF generation failed: {exc}", content_type="text/plain", status=500)

        # fallback: ensure a response always returned
        return self.render_to_response(ctx)

    # ---------- helper ----------
    def _build_entries(self, head):
        entries = []
        
        # helper: safely get firm label from different models
        def firm_label_from(obj):
            # try firm_name field
            name = getattr(obj, "firm_name", "") or ""
            firm_obj = getattr(obj, "firm", None)
            if not name and firm_obj:
                name = getattr(firm_obj, "firmname", str(firm_obj))
            return name or ""
        
        if getattr(head, "openingdebit", None) and head.openingdebit != Decimal("0"):
            entries.append({"entry_no": "OPEN", 
                            "date": None,
                            "debit": head.openingdebit,
                            "credit": Decimal("0"),
                            "remark": "Opening (Dr)",
                            "firm_name": "",
                        })
        elif getattr(head, "openingcredit", None) and head.openingcredit != Decimal("0"):
            entries.append({"entry_no": "OPEN", 
                            "date": None,
                            "debit": Decimal("0"), 
                            "credit": head.openingcredit,
                            "remark": "Opening (Cr)",
                            "firm_name": "",
                        })

        for s in SaleMaster.objects.filter(party=head).order_by("invdate"):
            entries.append({"entry_no": s.invno, 
                            "date": s.invdate,
                            "debit": s.netamt, 
                            "credit": Decimal("0"),
                            "remark": s.remark or f"Sale Inv#{s.invno}",
                            "firm_name": firm_label_from(s),
                           })
        for p in PurchaseMaster.objects.filter(party=head).order_by("invdate"):
            entries.append({"entry_no": p.invno,
                            "date": p.invdate,
                            "debit": Decimal("0"),
                            "credit": p.netamt,
                            "remark": p.remark or f"Purchase Inv#{p.invno}",
                            "firm_name": firm_label_from(p),
                        })
        for n in NaameEntry.objects.filter(party=head).order_by('daily_page__date'):
            entries.append({"entry_no": n.entry_no,
                            "date": n.daily_page.date,
                            "debit": n.amount, 
                            "credit": Decimal("0"),
                            "remark": n.remark or "Naame",
                            "firm_name": firm_label_from(n),
                          })
        for j in JamaEntry.objects.filter(party=head).order_by('daily_page__date'):
            entries.append({"entry_no": j.entry_no, 
                            "date": j.daily_page.date,
                            "debit": Decimal("0"), 
                            "credit": j.amount,
                            "remark": j.remark or "Jama",
                            "firm_name": firm_label_from(j),
                        })
        # sort + totals + running balance
        entries = sorted(entries, key=lambda x: (x["date"] is None, x["date"] or ""))
        total_debit = sum(e["debit"] for e in entries)
        total_credit = sum(e["credit"] for e in entries)
        bal = Decimal("0")
        for e in entries:
            bal += (e["debit"] or Decimal("0")) - (e["credit"] or Decimal("0"))
            e["balance"] = bal
        balance = total_debit - total_credit
        return entries, total_debit, total_credit, balance
    
class BrokerStatementView(TemplateView):
    """
    Single-URL Broker Statement view. POST name="action":
      - statement
      - print
      - export_excel
      - pdf
    """
    template_name = "brokerapp/account/broker_statement.html"
    printable_template = "brokerapp/account/broker_statement_printable.html"

    def get(self, request, *args, **kwargs):
        brokers = Broker.objects.order_by("brokername")
        ctx = {
            "brokers": brokers,
            "selected": None,
            "entries": [],
            "total_debit": Decimal("0"),
            "total_credit": Decimal("0"),
            "balance": Decimal("0"),
        }
        return self.render_to_response(ctx)

    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        broker_id = request.POST.get("broker") or request.GET.get("broker")
        brokers = Broker.objects.order_by("brokername")

        if not broker_id:
            ctx = {
                "brokers": brokers,
                "selected": None,
                "entries": [],
                "total_debit": Decimal("0"),
                "total_credit": Decimal("0"),
                "balance": Decimal("0"),
            }
            return self.render_to_response(ctx)

        selected = get_object_or_404(Broker, pk=broker_id)
        entries, total_debit, total_credit, balance = self._build_entries(selected)

        ctx = {
            "brokers": brokers,
            "selected": selected,
            "entries": entries,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "balance": balance,
            "today": date.today(),
        }

        # show statement in page
        if action in (None, "statement"):
            return self.render_to_response(ctx)

        # printable HTML
        if action == "print":
            return render(request, self.printable_template, ctx)

        # excel export
        if action == "export_excel":
            if Workbook is None:
                return HttpResponse(
                    "Required package 'openpyxl' not installed. Install with: pip install openpyxl",
                    content_type="text/plain",
                    status=500
                )
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Broker Statement"
                headers = ["Entry No", "Date", "Debit", "Credit", "Remark", "Balance"]
                ws.append(headers)
                for e in entries:
                    ws.append([
                        e.get("entry_no"),
                        e["date"].strftime("%Y-%m-%d") if e["date"] else "",
                        float(e.get("debit") or 0),
                        float(e.get("credit") or 0),
                        e.get("remark") or "",
                        float(e.get("balance") or 0),
                    ])
                ws.append([])
                ws.append(["", "Total", float(total_debit), float(total_credit), "", float(balance)])

                if get_column_letter:
                    for i, col in enumerate(ws.columns, start=1):
                        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                        ws.column_dimensions[get_column_letter(i)].width = max_len + 2

                out = io.BytesIO()
                wb.save(out)
                out.seek(0)
                resp = HttpResponse(
                    out.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                safe_name = "".join(ch if ord(ch) < 128 else "?" for ch in selected.brokername)[:40]
                resp["Content-Disposition"] = f'attachment; filename="broker_statement_{safe_name}.xlsx"'
                return resp
            except Exception as exc:
                return HttpResponse(f"Excel export failed: {exc}", content_type="text/plain", status=500)

        # pdf export
        if action == "pdf":
            if FPDF is None:
                return HttpResponse(
                    "Required package 'fpdf' not installed. Install with: pip install fpdf",
                    content_type="text/plain",
                    status=500
                )

            def safe_text(val, maxlen=None):
                s = "" if val is None else str(val)
                s = s.replace("—", "-").replace("–", "-")
                s = "".join(ch if ord(ch) < 128 else "?" for ch in s)
                return s[:maxlen] if maxlen else s

            try:
                pdf = FPDF()
                pdf.add_page()
                pdf.set_auto_page_break(auto=True, margin=10)

                pdf.set_font("Helvetica", "B", 14)
                pdf.cell(0, 10, safe_text(f"Broker Statement - {selected.brokername}", 140), ln=True, align="C")
                pdf.set_font("Helvetica", "", 10)
                pdf.cell(0, 6, safe_text(f"Generated on: {date.today().strftime('%d-%m-%Y')}", 80), ln=True, align="C")
                pdf.ln(4)

                headers = ["Entry No", "Date", "Debit", "Credit", "Remark", "Balance"]
                widths = [22, 22, 28, 28, 60, 30]
                pdf.set_font("Helvetica", "B", 9)
                for i, h in enumerate(headers):
                    pdf.cell(widths[i], 8, safe_text(h, 40), border=1, align="C")
                pdf.ln(8)

                pdf.set_font("Helvetica", "", 9)
                for e in entries:
                    vals = [
                        safe_text(e.get("entry_no", ""), 20),
                        safe_text(e["date"].strftime("%Y-%m-%d") if e["date"] else "", 20),
                        safe_text(f"{(e.get('debit') or Decimal('0')):.2f}", 20),
                        safe_text(f"{(e.get('credit') or Decimal('0')):.2f}", 20),
                        safe_text(e.get("remark", ""), 120),
                        safe_text(f"{(e.get('balance') or Decimal('0')):.2f}", 20),
                    ]
                    for i, v in enumerate(vals):
                        pdf.cell(widths[i], 7, v, border=1, align="L" if i in (0, 1, 4) else "R")
                    pdf.ln(7)

                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(widths[0] + widths[1], 8, safe_text("TOTAL", 40), border=1, align="L")
                pdf.cell(widths[2], 8, safe_text(f"{total_debit:.2f}", 20), border=1, align="R")
                pdf.cell(widths[3], 8, safe_text(f"{total_credit:.2f}", 20), border=1, align="R")
                pdf.cell(widths[4], 8, "", border=1, align="R")
                pdf.cell(widths[5], 8, safe_text(f"{balance:.2f}", 20), border=1, align="R")

                buf = io.BytesIO()
                pdf.output(buf)
                buf.seek(0)
                resp = HttpResponse(buf.read(), content_type="application/pdf")
                safe_name = "".join(ch if ord(ch) < 128 else "?" for ch in selected.brokername)[:40]
                resp["Content-Disposition"] = f'attachment; filename="broker_statement_{safe_name}.pdf"'
                return resp
            except Exception as exc:
                return HttpResponse(f"PDF generation failed: {exc}", content_type="text/plain", status=500)

        # fallback
        return self.render_to_response(ctx)

    def _build_entries(self, selected):
   
        entries = []

        # helper to decide order_by field name for a model
        def _order_field(model_cls, preferred):
            # return preferred if model has it, otherwise fallback to 'id'
            return preferred if hasattr(model_cls, preferred) else 'id'

        # 1) JamaEntry -> credit
        jama_order = _order_field(JamaEntry, 'created_at')
        jama_qs = JamaEntry.objects.filter(broker=selected).order_by(jama_order)
        for j in jama_qs:
            date_val = j.created_at.date() if getattr(j, 'created_at', None) else None
            amt = Decimal(str(j.amount or 0))
            entries.append({
                "entry_no": f"J-{j.entry_no}",
                "date": date_val,
                "debit": Decimal("0"),
                "credit": amt,
                "remark": (j.remark or "") + " (Jama)",
            })

        # 2) NaameEntry -> debit
        naame_order = _order_field(NaameEntry, 'created_at')
        naame_qs = NaameEntry.objects.filter(broker=selected).order_by(naame_order)
        for n in naame_qs:
            date_val = n.created_at.date() if getattr(n, 'created_at', None) else None
            amt = Decimal(str(n.amount or 0))
            entries.append({
                "entry_no": f"N-{n.entry_no}",
                "date": date_val,
                "debit": amt,
                "credit": Decimal("0"),
                "remark": (n.remark or "") + " (Naame)",
            })

        # 3) SaleMaster -> debit (using dramt)
        sale_order = _order_field(SaleMaster, 'invdate')
        sale_qs = SaleMaster.objects.filter(broker=selected).order_by(sale_order)
        for s in sale_qs:
            date_val = getattr(s, "invdate", None)
            amt = Decimal(str(getattr(s, "dramt", 0) or 0))
            entries.append({
                "entry_no": f"S-{getattr(s, 'invno', '')}",
                "date": date_val,
                "debit": amt,
                "credit": Decimal("0"),
                "remark": (getattr(s, "remark", "") or "") + " (Sale)",
            })

        # 4) PurchaseMaster -> credit (using dramt)
        purchase_order = _order_field(PurchaseMaster, 'invdate')
        purchase_qs = PurchaseMaster.objects.filter(broker=selected).order_by(purchase_order)
        for p in purchase_qs:
            date_val = getattr(p, "invdate", None)
            amt = Decimal(str(getattr(p, "dramt", 0) or 0))
            entries.append({
                "entry_no": f"P-{getattr(p, 'invno', '')}",
                "date": date_val,
                "debit": Decimal("0"),
                "credit": amt,
                "remark": (getattr(p, "remark", "") or "") + " (Purchase)",
            })

        # sort entries by date (None considered after real dates), then entry_no
        # Use a stable key: (is_date_none, date_or_max, entry_no)
        from datetime import datetime
        entries.sort(key=lambda x: (x["date"] is None, x["date"] or datetime.max.date(), x.get("entry_no", "")))

        # totals + running balance (Decimal)
        total_debit = sum(e["debit"] for e in entries) if entries else Decimal("0")
        total_credit = sum(e["credit"] for e in entries) if entries else Decimal("0")
        bal = Decimal("0")
        for e in entries:
            bal += (e["debit"] - e["credit"])
            e["balance"] = bal

        balance = bal
        return entries, total_debit, total_credit, balance


# --- AllBrokerBalanceView ---
class AllBrokerBalanceView(TemplateView):
    """
    Broker version of All Party Balance.
    Supports POST actions via buttons with name="action":
      - balance       : show table in page
      - print         : render printable HTML (user can browser-print)
      - export_excel  : return .xlsx (requires openpyxl)
      - pdf           : return PDF generated with fpdf2 (if installed)
    """
    template_name = "brokerapp/account/all_broker_balance.html"
    printable_template = "brokerapp/account/all_broker_balance_printable.html"

    # ---------- helpers ----------
    def _org_filter(self, qs):
        org_id = self.request.session.get("org_id")
        if not org_id:
            return qs
        field_names = [f.attname for f in qs.model._meta.fields]
        if "org_id" in field_names:
            return qs.filter(org_id=org_id)
        return qs

    def _sum(self, qs, field):
        """Safe sum returning Decimal(0) when None."""
        return qs.aggregate(t=Sum(field))["t"] or Decimal("0")

    # ---------- GET ----------
    def get(self, request, *args, **kwargs):
        today = date.today()
        ctx = self._build_context(start=today, end=today, broker=None)
        ctx["show_table"] = False
        return self.render_to_response(ctx)

    # ---------- POST ----------
    def post(self, request, *args, **kwargs):
        action = request.POST.get("action")
        today = date.today()

        # Build rows/totals (same data used by all actions)
        ctx = self._build_context(start=today, end=today, broker=None)

        # Balance -> show table in same template
        if action == "balance" or not action:
            ctx["show_table"] = True
            return self.render_to_response(ctx)

        # Print -> render printable HTML (no buttons)
        if action == "print":
            ctx["show_table"] = True
            return render(request, self.printable_template, ctx)

        # Export Excel -> create .xlsx (requires openpyxl)
        if action == "export_excel":
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except Exception:
                return HttpResponse(
                    "Required package 'openpyxl' not installed. Install with: pip install openpyxl",
                    content_type="text/plain",
                    status=500
                )

            wb = Workbook()
            ws = wb.active
            ws.title = "All Broker Balance"

            headers = ["Broker", "Op Dr", "Op Cr", "Opening", "Sale", "Purchase", "Naame", "Jama", "Balance"]
            ws.append(headers)

            for r in ctx["rows"]:
                bname = getattr(r["broker"], "brokername", str(r["broker"]))
                ws.append([
                    bname,
                    float(r["op_dr"]), float(r["op_cr"]),
                    float(r["opening"]), float(r["sale"]),
                    float(r["purchase"]), float(r["naame"]),
                    float(r["jama"]), float(r["balance"])
                ])

            # auto column width (simple)
            for i, col in enumerate(ws.columns, start=1):
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            resp = HttpResponse(
                out.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            resp["Content-Disposition"] = f'attachment; filename="all_broker_balance_{today}.xlsx"'
            return resp

        # PDF -> generate using fpdf2
        if action == "pdf":
            if FPDF is None:
                return HttpResponse(
                    "Required package 'fpdf2' not installed. Install with: pip install fpdf2",
                    content_type="text/plain",
                    status=500
                )

            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=10)

            # Header
            pdf.set_font("Helvetica", "B", 14)
            pdf.cell(0, 10, "All Broker Balance", ln=True, align="C")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 6, f"Generated on: {today.strftime('%d-%m-%Y')}", ln=True, align="C")
            pdf.ln(4)

            # Table headers
            headers = ["Broker", "Op Dr", "Op Cr", "Opening", "Sale", "Purchase", "Naame", "Jama", "Balance"]
            col_widths = [50, 18, 18, 24, 18, 22, 18, 18, 22]

            pdf.set_font("Helvetica", "B", 9)
            for i, h in enumerate(headers):
                pdf.cell(col_widths[i], 8, h, border=1, align="C")
            pdf.ln(8)

            # Rows
            pdf.set_font("Helvetica", "", 9)
            for r in ctx["rows"]:
                vals = [
                    getattr(r["broker"], "brokername", str(r["broker"])),
                    f"{r['op_dr']:.2f}", f"{r['op_cr']:.2f}",
                    f"{r['opening']:.2f}", f"{r['sale']:.2f}",
                    f"{r['purchase']:.2f}", f"{r['naame']:.2f}",
                    f"{r['jama']:.2f}", f"{r['balance']:.2f}"
                ]
                for i, v in enumerate(vals):
                    align = "L" if i == 0 else "R"
                    pdf.cell(col_widths[i], 7, v, border=1, align=align)
                pdf.ln(7)

            # Totals row
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(col_widths[0], 8, "TOTAL", border=1, align="L")
            totals = ctx["totals"]
            # print totals aligned with numeric columns (skip Opening total)
            pdf.cell(col_widths[1], 8, f"{totals['opdr']:.2f}", border=1, align="R")
            pdf.cell(col_widths[2], 8, f"{totals['opcr']:.2f}", border=1, align="R")
            pdf.cell(col_widths[3], 8, "", border=1, align="R")
            pdf.cell(col_widths[4], 8, f"{totals['sale']:.2f}", border=1, align="R")
            pdf.cell(col_widths[5], 8, f"{totals['purchase']:.2f}", border=1, align="R")
            pdf.cell(col_widths[6], 8, f"{totals['naame']:.2f}", border=1, align="R")
            pdf.cell(col_widths[7], 8, f"{totals['jama']:.2f}", border=1, align="R")
            pdf.cell(col_widths[8], 8, f"{totals['balance']:.2f}", border=1, align="R")
            pdf.ln(10)

            buf = io.BytesIO()
            pdf.output(buf)
            buf.seek(0)
            resp = HttpResponse(buf.read(), content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="all_broker_balance_{today}.pdf"'
            return resp

        # Unknown action -> render without table
        ctx["show_table"] = False
        return self.render_to_response(ctx)

    # ---------- core calculation ----------
    def _build_context(self, start, end, broker):
        # brokers
        from .models import Broker, HeadParty, SaleMaster, PurchaseMaster, NaameEntry, JamaEntry, DailyPage

        brokers = Broker.objects.all().order_by("brokername")
        if broker:
            brokers = brokers.filter(brokername=broker.brokername)

        rows = []
        totals = {
            "opdr": Decimal("0"), "opcr": Decimal("0"),
            "sale": Decimal("0"), "purchase": Decimal("0"),
            "naame": Decimal("0"), "jama": Decimal("0"),
            "balance": Decimal("0")
        }

        dp_before = DailyPage.objects.filter(date__lt=start)
        dp_range = DailyPage.objects.filter(date__range=(start, end))

        org_id = self.request.session.get("org_id")
        if org_id:
            dp_before = dp_before.filter(org_id=org_id)
            dp_range = dp_range.filter(org_id=org_id)
            brokers = brokers.filter(org_id=org_id)

        for b in brokers:
            op_dr = Decimal(getattr(b, "openingdebit", 0) or 0)
            op_cr = Decimal(getattr(b, "openingcredit", 0) or 0)

            sale_before = self._sum(self._org_filter(SaleMaster.objects.filter(broker=b, invdate__lt=start)), "netamt")
            purch_before = self._sum(self._org_filter(PurchaseMaster.objects.filter(broker=b, invdate__lt=start)), "netamt")
            naame_before = self._sum(NaameEntry.objects.filter(daily_page__in=dp_before, broker=b), "amount")
            jama_before = self._sum(JamaEntry.objects.filter(daily_page__in=dp_before, broker=b), "amount")

            opening = (op_dr - op_cr) + (sale_before - purch_before + naame_before - jama_before)

            sale = self._sum(self._org_filter(SaleMaster.objects.filter(broker=b, invdate__range=(start, end))), "netamt")
            purchase = self._sum(self._org_filter(PurchaseMaster.objects.filter(broker=b, invdate__range=(start, end))), "netamt")
            naame = self._sum(NaameEntry.objects.filter(daily_page__in=dp_range, broker=b), "amount")
            jama = self._sum(JamaEntry.objects.filter(daily_page__in=dp_range, broker=b), "amount")

            balance = opening + sale - purchase + naame - jama

            rows.append({
                "broker": b, "op_dr": op_dr, "op_cr": op_cr, "opening": opening,
                "sale": sale, "purchase": purchase, "naame": naame,
                "jama": jama, "balance": balance
            })

            totals["opdr"] += op_dr
            totals["opcr"] += op_cr
            totals["sale"] += sale
            totals["purchase"] += purchase
            totals["naame"] += naame
            totals["jama"] += jama
            totals["balance"] += balance

        return {"rows": rows, "totals": totals, "start": start, "end": end}
